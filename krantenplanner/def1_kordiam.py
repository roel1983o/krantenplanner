# Auto-generated from DEF1 - Kordiam Parser.ipynb
from __future__ import annotations

import os
import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

TARGET_SHEET = "Totale verhalenlijst"
SOURCE_SHEET = "Story List"

SHEETS_TO_REMOVE = ["Statistics", "Aggregated story list"]

# Headers in A1.. (dynamisch op basis van deze lijst)
# NB: 'Classificatie' hoort tussen Focusregio en Heel Limburg.
TARGET_HEADERS: List[str] = ["Story ID", "Naam productie", "Note", "Publ. status", "Focusregio", "Classificatie", "Heel Limburg", "Voorkeurspositie", "Beeld voor print", "Publicatiedwang", "Top 8", "Karakters", "Artikelsoort", "Leverancier", "Auteur", "Gewenste placeholder", "Tweede keus placeholder", "Derde keus placeholder", "Vierde keus placeholder", "Placeholder bij enigszins geschikt", "Prioscore", "Gekozen template", "Gekozen placeholder", "Placeholder-concessie", "Plaatsing"]

# Mapping: doelkolom -> bronkolom (op headernaam). Niet-gemapte kolommen blijven leeg.
COL_MAP: Dict[str, str] = {
    "Story ID": "Story ID",
    "Naam productie": "Description",
    "Note": "Note",
    "Publ. status": "Publ. status",
    "Focusregio": "Focusregio",
    "Classificatie": "Classificatie",
    "Heel Limburg": "Heel Limburg",
    "Voorkeurspositie": "Voorkeurspositie",
    "Beeld voor print": "Beeld voor print",
    "Publicatiedwang": "Publicatiedwang",
    "Top 8": "Top 8",
    "Karakters": "Text length"
}

# Fixed text in AA1 en AB1 (geen headers)
FIXED_TEXT = "Totale verhalenlijst"

def find_header_row(ws, needle: str = "Story ID", max_scan_rows: int = 200) -> int:
    for r in range(1, max_scan_rows + 1):
        row_vals = [c.value for c in ws[r]]
        if needle in row_vals:
            return r
    raise ValueError(f"Kon header-rij niet vinden: '{needle}' niet gevonden in eerste {max_scan_rows} rijen.")


def build_header_index(ws, header_row: int) -> Dict[str, int]:
    idx: Dict[str, int] = {}
    for col_i, cell in enumerate(ws[header_row], start=1):
        val = cell.value
        if isinstance(val, str) and val.strip():
            idx[val.strip()] = col_i
    return idx


def clear_or_create_sheet(wb: openpyxl.Workbook, name: str):
    if name in wb.sheetnames:
        wb.remove(wb[name])
    return wb.create_sheet(title=name)


def write_target_headers(ws_target, headers: List[str]) -> None:
    # Opmaak: Bold + 20% grijs
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for i, h in enumerate(headers, start=1):
        cell = ws_target.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill


def set_fixed_text(ws_target) -> None:
    ws_target.cell(row=1, column=27, value=FIXED_TEXT)  # AA1
    ws_target.cell(row=1, column=28, value=FIXED_TEXT)  # AB1


def iter_data_rows(ws_source, header_row: int):
    r = header_row + 1
    while True:
        row = ws_source[r]
        values = [c.value for c in row]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
            break
        yield r
        r += 1


def process_kordiam(input_xlsx: str, output_xlsx: str, mapping_xlsx: str = "") -> Tuple[int, List[str]]:
    if not os.path.exists(input_xlsx):
        raise FileNotFoundError(f"Inputbestand niet gevonden: {input_xlsx}")

    wb = openpyxl.load_workbook(input_xlsx)

    # v26: parse planning date from 'Story List'!A3 (last 10 chars)
    planning_date = None
    try:
        if SOURCE_SHEET in wb.sheetnames:
            raw_a3 = wb[SOURCE_SHEET]["A3"].value
            if raw_a3 is not None:
                s = str(raw_a3)
                s10 = s[-10:].replace(".", "-")
                for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
                    try:
                        planning_date = datetime.datetime.strptime(s10, fmt).date()
                        break
                    except Exception:
                        pass
    except Exception:
        planning_date = None


    # 1) remove sheets
    for s in SHEETS_TO_REMOVE:
        if s in wb.sheetnames:
            wb.remove(wb[s])

    if SOURCE_SHEET not in wb.sheetnames:
        raise ValueError(f"Bron-sheet '{SOURCE_SHEET}' ontbreekt. Aanwezige sheets: {wb.sheetnames}")
    ws_source = wb[SOURCE_SHEET]

    # 2) create target sheet
    ws_target = clear_or_create_sheet(wb, TARGET_SHEET)
    write_target_headers(ws_target, TARGET_HEADERS)

    # 3) AA1/AB1 fixed text
    set_fixed_text(ws_target)

    header_row = find_header_row(ws_source, needle="Story ID")
    src_idx = build_header_index(ws_source, header_row)

    warnings: List[str] = []
    if mapping_xlsx and not os.path.exists(mapping_xlsx):
        warnings.append(f"WAARSCHUWING: mappingbestand niet gevonden: {mapping_xlsx} (stap 9 wordt overgeslagen)")
    elif not mapping_xlsx:
        warnings.append("WAARSCHUWING: mappingbestand (MAPPING_XLSX) is leeg/niet gezet (stap 9 wordt overgeslagen)")

    required_sources = set(COL_MAP.values()) | {"Assignee last name", "Assignee first name", "Type verhaal", "Group"}
    missing = sorted({src for src in required_sources if src not in src_idx})
    for m in missing:
        warnings.append(f"WAARSCHUWING: bronkolom ontbreekt in '{SOURCE_SHEET}': {m} (kolom wordt leeg gelaten)")

    # 9) Placeholder-mapping voorbereiden
    placeholder_lookup = {}
    if mapping_xlsx and os.path.exists(mapping_xlsx):
        try:
            wb_map = openpyxl.load_workbook(mapping_xlsx, data_only=True)
            ws_map = None
            for name in wb_map.sheetnames:
                ws_try = wb_map[name]
                header = [ws_try.cell(1, c).value for c in range(1, ws_try.max_column + 1)]
                header_norm = [h.strip() if isinstance(h, str) else h for h in header]
                if "Artikelsoort" in header_norm and "Beeld voor print" in header_norm and "Top 8" in header_norm:
                    ws_map = ws_try
                    break
            if ws_map is None:
                warnings.append(f"WAARSCHUWING: mappingbestand heeft geen herkenbare header (stap 9 wordt overgeslagen): {mapping_xlsx}")
            else:
                header = [ws_map.cell(1, c).value for c in range(1, ws_map.max_column + 1)]
                header_norm = [h.strip() if isinstance(h, str) else h for h in header]
                col_idx = {h: i + 1 for i, h in enumerate(header_norm) if isinstance(h, str) and h}

                def _norm(v):
                    if v is None:
                        return "(geen waarde ingevuld)"
                    if isinstance(v, str):
                        s = v.strip()
                        return s if s else "(geen waarde ingevuld)"
                    return str(v).strip()

                for rr in range(2, ws_map.max_row + 1):
                    artikel = ws_map.cell(rr, col_idx.get("Artikelsoort", 1)).value
                    if artikel is None or (isinstance(artikel, str) and not artikel.strip()):
                        break
                    beeld = ws_map.cell(rr, col_idx.get("Beeld voor print", 2)).value
                    top8 = ws_map.cell(rr, col_idx.get("Top 8", 3)).value
                    key = (_norm(artikel), _norm(beeld), _norm(top8))
                    placeholder_lookup[key] = {
                        "Gewenste placeholder": ws_map.cell(rr, col_idx.get("Gewenste placeholder", 0)).value if col_idx.get("Gewenste placeholder") else None,
                        "Tweede keus placeholder": ws_map.cell(rr, col_idx.get("Tweede keus placeholder", 0)).value if col_idx.get("Tweede keus placeholder") else None,
                        "Derde keus placeholder": ws_map.cell(rr, col_idx.get("Derde keus placeholder", 0)).value if col_idx.get("Derde keus placeholder") else None,
                        "Vierde keus placeholder": ws_map.cell(rr, col_idx.get("Vierde keus placeholder", 0)).value if col_idx.get("Vierde keus placeholder") else None,
                        "Placeholder bij enigszins geschikt": ws_map.cell(rr, col_idx.get("Placeholder bij enigszins geschikt", 0)).value if col_idx.get("Placeholder bij enigszins geschikt") else None,
                    }
        except Exception as e:
            warnings.append(f"WAARSCHUWING: kon mappingbestand niet lezen ({mapping_xlsx}); stap 9 wordt overgeslagen. Fout: {e}")

    unmapped_placeholder_rows = 0

    out_row = 2
    mapped_rows = 0
    removed_rows = 0

    for r in iter_data_rows(ws_source, header_row):
        story_id_col = src_idx.get("Story ID")
        story_id_val = ws_source.cell(row=r, column=story_id_col).value if story_id_col else None
        if story_id_val is None or (isinstance(story_id_val, str) and not story_id_val.strip()):
            break

        # 10) filter Column/Rubriek
        col_tv_filter = src_idx.get("Type verhaal")
        tv_filter_val = ws_source.cell(row=r, column=col_tv_filter).value if col_tv_filter is not None else None
        tv_filter_str = tv_filter_val.strip() if isinstance(tv_filter_val, str) else (str(tv_filter_val).strip() if tv_filter_val is not None else "")
        if tv_filter_str in {"Column", "Rubriek"}:
            removed_rows += 1
            continue

        # basis mapping
        for target_col_i, target_header in enumerate(TARGET_HEADERS, start=1):
            src_header = COL_MAP.get(target_header)
            if src_header is None:
                continue
            src_col_i = src_idx.get(src_header)
            if src_col_i is None:
                continue
            ws_target.cell(row=out_row, column=target_col_i, value=ws_source.cell(row=r, column=src_col_i).value)

        # 6) Auteur
        auteur_last = ws_source.cell(row=r, column=src_idx["Assignee last name"]).value if "Assignee last name" in src_idx else None
        auteur_first = ws_source.cell(row=r, column=src_idx["Assignee first name"]).value if "Assignee first name" in src_idx else None
        auteur_last = auteur_last.strip() if isinstance(auteur_last, str) else auteur_last
        auteur_first = auteur_first.strip() if isinstance(auteur_first, str) else auteur_first
        if auteur_last and auteur_first:
            auteur_val = f"{auteur_last}, {auteur_first}"
        elif auteur_last:
            auteur_val = str(auteur_last)
        elif auteur_first:
            auteur_val = str(auteur_first)
        else:
            auteur_val = None
        ws_target.cell(row=out_row, column=TARGET_HEADERS.index("Auteur") + 1, value=auteur_val)

        # 7) Artikelsoort
        text_len_val = ws_source.cell(row=r, column=src_idx["Text length"]).value if "Text length" in src_idx else None
        type_verhaal_val = ws_source.cell(row=r, column=src_idx["Type verhaal"]).value if "Type verhaal" in src_idx else None
        try:
            tl_int = int(float(str(text_len_val).strip())) if text_len_val is not None and str(text_len_val).strip() != "" else None
        except Exception:
            tl_int = None
        tv_str = str(type_verhaal_val).strip() if type_verhaal_val is not None else ""
        artikelsoort = None
        if tl_int == 7200:
            artikelsoort = "XXL"
        elif tl_int == 5400:
            artikelsoort = "XL"
        elif tl_int == 4000:
            artikelsoort = "L"
        elif tl_int == 2800:
            artikelsoort = "M_nws" if tv_str == "Nieuws" else "M_lk"
        elif tl_int == 1800:
            artikelsoort = "S_nws" if tv_str == "Nieuws" else "S_lk"
        elif tl_int == 1000:
            artikelsoort = "XS"
        ws_target.cell(row=out_row, column=TARGET_HEADERS.index("Artikelsoort") + 1, value=artikelsoort)

        # 8) Leverancier
        group_val = ws_source.cell(row=r, column=src_idx["Group"]).value if "Group" in src_idx else None
        supplier_map = {
            "Nieuwsdienst": "rND",
            "Maastricht - Heuvelland": "rMH",
            "Sittard-Geleen": "rSG",
            "Parkstad": "rPS",
            "Noord-Limburg": "rNO",
            "Midden-Limburg": "rMI",
            "Economie": "rEC",
            "Cultuur & Media": "rCU",
            "LS": "rLS",
            "Onderzoek": "rOZ",
            "Opinie": "rOP",
            "Sport": "rSP",
        }
        leverancier_val = supplier_map.get(str(group_val).strip() if group_val is not None else None)
        ws_target.cell(row=out_row, column=TARGET_HEADERS.index("Leverancier") + 1, value=leverancier_val)

        # 9) Placeholders
        if placeholder_lookup:
            def _norm(v):
                if v is None:
                    return "(geen waarde ingevuld)"
                if isinstance(v, str):
                    s = v.strip()
                    return s if s else "(geen waarde ingevuld)"
                return str(v).strip()

            beeld_val = ws_source.cell(row=r, column=src_idx["Beeld voor print"]).value if "Beeld voor print" in src_idx else None
            top8_val = ws_source.cell(row=r, column=src_idx["Top 8"]).value if "Top 8" in src_idx else None
            key = (_norm(artikelsoort), _norm(beeld_val), _norm(top8_val))
            rule = placeholder_lookup.get(key)
            if rule:
                for h, v in rule.items():
                    if h in TARGET_HEADERS:
                        ws_target.cell(row=out_row, column=TARGET_HEADERS.index(h) + 1, value=v)
            else:
                unmapped_placeholder_rows += 1

        mapped_rows += 1
        out_row += 1

    if placeholder_lookup and unmapped_placeholder_rows:
        warnings.append(f"WAARSCHUWING: stap 9 (placeholder mapping) geen match gevonden voor {unmapped_placeholder_rows} rij(en); placeholders zijn daar leeg gelaten.")
    if removed_rows:
        warnings.append(f"INFO: stap 10 (filter Column/Rubriek) verwijderd: {removed_rows} rij(en).")

    # 11) verwijder Story List
    if SOURCE_SHEET in wb.sheetnames:
        wb.remove(wb[SOURCE_SHEET])

    # Kandidatenlijsten maken
    candidate_names = ["GO-01", "GO-02", "NM-NO", "NM-MI", "ZU-SG", "ZU-PS", "ZU-MH"]
    for nm in candidate_names:
        if nm in wb.sheetnames:
            wb.remove(wb[nm])
    base_ws = wb[TARGET_SHEET]
    for nm in candidate_names:
        ws_copy = wb.copy_worksheet(base_ws)
        ws_copy.title = nm

    # Kandidatenlijsten maken: opmaak + fixed texts
    gray50 = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    thick_side = Side(style="thick")
    fixed_by_sheet = {
        "GO-01": ("rND; rOZ; rEC;", "Limburg-breed"),
        "GO-02": ("rND; rOZ; rEC;", "Limburg-breed"),
        "NM-NO": ("rNO", "Noord"),
        "NM-MI": ("rMI", "Midden"),
        "ZU-SG": ("rSG", "Sittard"),
        "ZU-PS": ("rPS", "Parkstad"),
        "ZU-MH": ("rMH", "Maastricht"),
    }

    def _apply_right_border(cell, side):
        b = cell.border
        cell.border = Border(
            left=b.left, right=side, top=b.top, bottom=b.bottom,
            diagonal=b.diagonal, diagonal_direction=b.diagonal_direction,
            outline=b.outline, vertical=b.vertical, horizontal=b.horizontal
        )

    for nm in candidate_names:
        ws_c = wb[nm]

        # Verticale lijn tussen Z en AA: dikke rechterborder op kolom Z
        max_r = max(ws_c.max_row, 40)
        for rr in range(1, max_r + 1):
            _apply_right_border(ws_c.cell(row=rr, column=26), thick_side)

        # AA1:AF40 50% grijs
        for rr in range(1, 41):
            for cc in range(27, 33):  # AA..AF
                ws_c.cell(row=rr, column=cc).fill = gray50

        # Fixed text AE1/AF1
        if nm in fixed_by_sheet:
            ae1, af1 = fixed_by_sheet[nm]
            ws_c.cell(row=1, column=31, value=ae1)  # AE1
            ws_c.cell(row=1, column=32, value=af1)  # AF1

        # Kandidatenlijsten voor regiospreads bewerken (NM-NO, NM-MI, ZU-SG, ZU-PS, ZU-MH)
        if nm in {"NM-NO", "NM-MI", "ZU-SG", "ZU-PS", "ZU-MH"}:
            focus_col = TARGET_HEADERS.index("Focusregio") + 1
            af1_val = ws_c.cell(row=1, column=32).value  # AF1
            af1_key = str(af1_val).strip().lower() if af1_val is not None else ""
            keep_alt = "limburg-breed"

            # 1) verwijderen: focus bevat niet ab2 én niet limburg-breed
            for rr in range(ws_c.max_row, 1, -1):
                focus_val = ws_c.cell(row=rr, column=focus_col).value
                focus_str = str(focus_val).strip().lower() if focus_val is not None else ""
                if (af1_key not in focus_str) and (keep_alt not in focus_str):
                    ws_c.delete_rows(rr, 1)

            # 2) verwijderen: focus bevat limburg-breed + extra waarde, maar niet ab2
            for rr in range(ws_c.max_row, 1, -1):
                focus_val = ws_c.cell(row=rr, column=focus_col).value
                focus_str = str(focus_val).strip().lower() if focus_val is not None else ""
                if keep_alt in focus_str and (af1_key not in focus_str):
                    norm = focus_str
                    for sep in [",", "\n", "/", "|"]:
                        norm = norm.replace(sep, ";")
                    parts = [p.strip() for p in norm.split(";") if p and p.strip()]
                    if ("limburg-breed" in parts) and (len(set(parts)) >= 2):
                        ws_c.delete_rows(rr, 1)

            # 3) classificatie instellen
            class_col = TARGET_HEADERS.index("Classificatie") + 1
            chars_col = TARGET_HEADERS.index("Karakters") + 1
            for rr in range(2, ws_c.max_row + 1):
                focus_val = ws_c.cell(row=rr, column=focus_col).value
                focus_str = str(focus_val).strip().lower() if focus_val is not None else ""
                chars_val = ws_c.cell(row=rr, column=chars_col).value
                try:
                    chars_num = int(float(str(chars_val).strip())) if chars_val is not None and str(chars_val).strip() != "" else None
                except Exception:
                    chars_num = None

                new_class = None
                if af1_key and (af1_key in focus_str):
                    new_class = "A-keus; B-keus; C-keus;"
                else:
                    if focus_str == "limburg-breed":
                        if chars_num is not None and chars_num > 4001:
                            new_class = "C-keus"
                        else:
                            new_class = "B-keus; C-keus"

                if new_class is not None:
                    ws_c.cell(row=rr, column=class_col, value=new_class)
            # 4) Prioscore berekenen (beginwaarde 0)
            prio_col = TARGET_HEADERS.index("Prioscore") + 1
            top8_col = TARGET_HEADERS.index("Top 8") + 1
            pubd_col = TARGET_HEADERS.index("Publicatiedwang") + 1
            hl_col = TARGET_HEADERS.index("Heel Limburg") + 1
            pref_col = TARGET_HEADERS.index("Voorkeurspositie") + 1
            lev_col = TARGET_HEADERS.index("Leverancier") + 1
            ae1_val = ws_c.cell(row=1, column=31).value  # AE1
            ae1_str = str(ae1_val).strip().lower() if ae1_val is not None else ""

            for rr in range(2, ws_c.max_row + 1):
                score = 0

                class_val = ws_c.cell(row=rr, column=class_col).value
                class_str = str(class_val).lower() if class_val is not None else ""

                top8_val = ws_c.cell(row=rr, column=top8_col).value
                top8_str = str(top8_val).strip().lower() if top8_val is not None else ""

                pubd_val = ws_c.cell(row=rr, column=pubd_col).value
                pubd_str = str(pubd_val).strip().lower() if pubd_val is not None else ""

                hl_val = ws_c.cell(row=rr, column=hl_col).value
                hl_str = str(hl_val).strip().lower() if hl_val is not None else ""

                pref_val = ws_c.cell(row=rr, column=pref_col).value

                if top8_str == "ja" and ("a-keus" in class_str):
                    score += 2
                if "a-keus" in class_str:
                    score += 5

                if hl_str == "moet mee" and ("a-keus" in class_str):
                    score += 3

                if pubd_str == "ja":
                    score += 3
                elif pubd_str == "nee":
                    score -= 1
                elif pubd_str == "op te sparen":
                    score += 1

                # Leverancier-regel: IF Leverancier = waarde cel AE1 THEN +1
                lev_val = ws_c.cell(row=rr, column=lev_col).value
                lev_str = str(lev_val).strip().lower() if lev_val is not None else ""
                if ae1_str and lev_str == ae1_str:
                    score += 1

                # Voorkeurspositie-regels t.o.v. naam tabblad (sheet)
                sheet_name = nm
                if pref_val is not None and str(pref_val).strip() != "":
                    pref_str = str(pref_val).strip()
                    if pref_str == sheet_name:
                        score += 20
                    else:
                        # v13 fix: alleen -20 als Voorkeurspositie niet 'Nee' is en niet de sheetnaam
                        if pref_str.lower() != "nee":
                            score -= 20

                ws_c.cell(row=rr, column=prio_col, value=score)

            # 5) Sorteer op Prioscore (hoog -> laag) binnen het tabblad (alleen data A..Y; fixed blok AA..AF blijft staan)
            data_cols = len(TARGET_HEADERS)  # A..Y
            rows_data = []
            for rr in range(2, ws_c.max_row + 1):
                row_vals = [ws_c.cell(row=rr, column=cc).value for cc in range(1, data_cols + 1)]
                prio_val = row_vals[prio_col - 1]
                try:
                    prio_num = float(prio_val)
                except Exception:
                    prio_num = float("-inf")
                rows_data.append((prio_num, row_vals))

            rows_data.sort(key=lambda x: x[0], reverse=True)

            for i, (_prio, row_vals) in enumerate(rows_data, start=2):
                for cc, v in enumerate(row_vals, start=1):
                    c = ws_c.cell(row=i, column=cc)
                    c.value = v


            # Re-apply grijs blok + fixed texts
            for rr in range(1, 41):
                for cc in range(27, 33):
                    ws_c.cell(row=rr, column=cc).fill = gray50
            # (v15+) AA2/AB2 worden niet meer gebruikt; leeg laten.
            ws_c.cell(row=2, column=27, value=None)
            ws_c.cell(row=2, column=28, value=None)



        # Kandidatenlijsten voor GO-spreads bewerken (GO-01, GO-02) (v17a)
        if nm in {"GO-01", "GO-02"}:
            heel_col = TARGET_HEADERS.index("Heel Limburg") + 1
            focus_col = TARGET_HEADERS.index("Focusregio") + 1
            kar_col = TARGET_HEADERS.index("Karakters") + 1
            class_col = TARGET_HEADERS.index("Classificatie") + 1

            # 1) Verwijder alle rijen waarvoor geldt: Heel Limburg≠geschikt AND Heel Limburg≠moet mee
            for rr in range(ws_c.max_row, 1, -1):
                heel_val = ws_c.cell(row=rr, column=heel_col).value
                heel_key = str(heel_val).strip().lower() if heel_val is not None else ""
                if heel_key not in {"geschikt", "moet mee"}:
                    ws_c.delete_rows(rr, 1)

            # 2) Bepaal per rij de waarde voor Classificatie (GO-regels)
            leverancier_col = TARGET_HEADERS.index("Leverancier") + 1
            for rr in range(2, ws_c.max_row + 1):
                lev_val = ws_c.cell(row=rr, column=leverancier_col).value
                lev_key = str(lev_val).strip().lower() if lev_val is not None else ""

                focus_val = ws_c.cell(row=rr, column=focus_col).value
                focus_key = str(focus_val).strip().lower() if focus_val is not None else ""

                kar_val = ws_c.cell(row=rr, column=kar_col).value
                try:
                    kar_num = int(float(kar_val)) if kar_val is not None and str(kar_val).strip() != "" else None
                except Exception:
                    kar_num = None

                # v18:
                # IF Leverancier=rND OR rOZ OR rEC THEN A/B/C
                # IF Focusregio CONTAINS 'Limburg-breed' THEN A/B/C
                # ELSE B/C; and if Karakters>4001 THEN C
                if lev_key in {"rnd", "roz", "rec"}:
                    class_val = "A-keus; B-keus; C-keus;"
                elif "limburg-breed" in focus_key:
                    class_val = "A-keus; B-keus; C-keus;"
                else:
                    class_val = "B-keus; C-keus;"
                    if kar_num is not None and kar_num > 4001:
                        class_val = "C-keus;"

                ws_c.cell(row=rr, column=class_col, value=class_val)

            # 4) Prioscore berekenen (beginwaarde 0) (GO-01/GO-02 v19)
            prio_col = TARGET_HEADERS.index("Prioscore") + 1
            top8_col = TARGET_HEADERS.index("Top 8") + 1
            pubd_col = TARGET_HEADERS.index("Publicatiedwang") + 1
            hl_col = TARGET_HEADERS.index("Heel Limburg") + 1
            pref_col = TARGET_HEADERS.index("Voorkeurspositie") + 1
            lev_col = TARGET_HEADERS.index("Leverancier") + 1

            for rr in range(2, ws_c.max_row + 1):
                score = 0

                class_val = ws_c.cell(row=rr, column=class_col).value
                class_str = str(class_val).lower() if class_val is not None else ""

                top8_val = ws_c.cell(row=rr, column=top8_col).value
                top8_str = str(top8_val).strip().lower() if top8_val is not None else ""

                pubd_val = ws_c.cell(row=rr, column=pubd_col).value
                pubd_str = str(pubd_val).strip().lower() if pubd_val is not None else ""

                hl_val = ws_c.cell(row=rr, column=hl_col).value
                hl_str = str(hl_val).strip().lower() if hl_val is not None else ""

                pref_val = ws_c.cell(row=rr, column=pref_col).value

                if top8_str == "ja" and ("a-keus" in class_str):
                    score += 2
                if "a-keus" in class_str:
                    score += 5

                if hl_str == "moet mee" and ("a-keus" in class_str):
                    score += 3

                if pubd_str == "ja":
                    score += 3
                elif pubd_str == "nee":
                    score -= 1
                elif pubd_str == "op te sparen":
                    score += 1

                # Leverancier-regel GO: IF Leverancier=rND OR rOZ OR rEC THEN +1
                lev_val = ws_c.cell(row=rr, column=lev_col).value
                lev_str = str(lev_val).strip().lower() if lev_val is not None else ""
                if lev_str in {"rnd", "roz", "rec"}:
                    score += 1

                # Voorkeurspositie-regels t.o.v. naam tabblad (sheet)
                sheet_name = nm
                if pref_val is not None and str(pref_val).strip() != "":
                    pref_str = str(pref_val).strip()
                    if pref_str == sheet_name:
                        score += 20
                    else:
                        if pref_str.lower() != "nee":
                            score -= 20

                ws_c.cell(row=rr, column=prio_col, value=score)

            # 5) Sorteer op Prioscore (hoog -> laag) binnen het tabblad (alleen data A..Y; fixed blok AA..AF blijft staan)
            data_cols = len(TARGET_HEADERS)  # A..Y
            rows_data = []
            for rr in range(2, ws_c.max_row + 1):
                row_vals = [ws_c.cell(row=rr, column=cc).value for cc in range(1, data_cols + 1)]
                prio_val = row_vals[prio_col - 1]
                try:
                    prio_num = float(prio_val)
                except Exception:
                    prio_num = float("-inf")
                rows_data.append((prio_num, row_vals))

            rows_data.sort(key=lambda x: x[0], reverse=True)

            for i, (_prio, row_vals) in enumerate(rows_data, start=2):
                for cc, v in enumerate(row_vals, start=1):
                    c = ws_c.cell(row=i, column=cc)
                    c.value = v


    # === Tabblad Stats maken en voorbereiden (v16) ===
    stats_name = "Stats"
    if stats_name in wb.sheetnames:
        wb.remove(wb[stats_name])
    ws_stats = wb.create_sheet(title=stats_name)

    # v26: planning date in Stats!AA1
    if planning_date is not None:
        cdt = ws_stats["AA1"]
        cdt.value = planning_date
        cdt.number_format = "yyyy-mm-dd"

    stats_headers = ["Tabblad", "St_aantal", "St_karakters", "St_fotos_dr", "St_fotos_overig", "St_fotos_flex", "Complexiteit", "Akpp_ondergrens", "Akpp_range", "Akpp_range_boost"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for ci, h in enumerate(stats_headers, start=1):
        cell = ws_stats.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill

    ordered_tabs = ["GO-01", "GO-02", "NM-NO", "NM-MI", "ZU-SG", "ZU-PS", "ZU-MH"]
    ordered_tabs = [t for t in ordered_tabs if t in wb.sheetnames]

    complexity_map = {}
    akpp_range_map = {}
    akpp_range_boost_map = {}

    def _get_col(ws, header_name: str):
        max_c = ws.max_column
        headers = [ws.cell(1, c).value for c in range(1, max_c + 1)]
        try:
            return headers.index(header_name) + 1
        except ValueError:
            return None

    def _count_a_keus(ws):
        class_col = _get_col(ws, "Classificatie")
        if not class_col:
            return 0
        cnt = 0
        target = "A-keus; B-keus; C-keus;"
        for rr in range(2, ws.max_row + 1):
            v = ws.cell(rr, class_col).value
            if v == target:
                cnt += 1
        return cnt

    def _sum_karakters_a_keus(ws):
        class_col = _get_col(ws, "Classificatie")
        kar_col = _get_col(ws, "Karakters")
        if not class_col or not kar_col:
            return 0
        target = "A-keus; B-keus; C-keus;"
        total = 0
        for rr in range(2, ws.max_row + 1):
            if ws.cell(rr, class_col).value == target:
                kv = ws.cell(rr, kar_col).value
                try:
                    total += int(kv) if kv is not None and str(kv).strip() != "" else 0
                except Exception:
                    # als Karakters geen integer is, tel dan 0 mee
                    total += 0
        return total

    def _count_fotos_dr_a_keus(ws):
        class_col = _get_col(ws, "Classificatie")
        beeld_col = _get_col(ws, "Beeld voor print")
        if not class_col or not beeld_col:
            return 0
        target = "A-keus; B-keus; C-keus;"
        dr_values = {"Dragend en bijplaat", "Dragend of bijplaat", "Dragend", "Flexibel"}
        cnt = 0
        for rr in range(2, ws.max_row + 1):
            if ws.cell(rr, class_col).value == target:
                bv = ws.cell(rr, beeld_col).value
                if bv in dr_values:
                    cnt += 1
        return cnt

    def _count_fotos_overig_a_keus(ws):
        class_col = _get_col(ws, "Classificatie")
        beeld_col = _get_col(ws, "Beeld voor print")
        if not class_col or not beeld_col:
            return 0
        target = "A-keus; B-keus; C-keus;"
        cnt = 0
        for rr in range(2, ws.max_row + 1):
            if ws.cell(rr, class_col).value == target:
                bv = ws.cell(rr, beeld_col).value
                if bv is None or str(bv).strip() == "" or bv == "Bijplaat":
                    cnt += 1
        return cnt

    def _count_fotos_flex_a_keus(ws):
        class_col = _get_col(ws, "Classificatie")
        beeld_col = _get_col(ws, "Beeld voor print")
        if not class_col or not beeld_col:
            return 0
        target = "A-keus; B-keus; C-keus;"
        cnt = 0
        for rr in range(2, ws.max_row + 1):
            if ws.cell(rr, class_col).value == target:
                bv = ws.cell(rr, beeld_col).value
                if bv == "Flexibel":
                    cnt += 1
        return cnt



    for ri, tab in enumerate(ordered_tabs, start=2):
        ws_tab = wb[tab]
        ws_stats.cell(row=ri, column=1, value=tab)

        st_aantal = _count_a_keus(ws_tab)
        st_kar = _sum_karakters_a_keus(ws_tab)
        st_fotos_dr = _count_fotos_dr_a_keus(ws_tab)
        st_fotos_overig = _count_fotos_overig_a_keus(ws_tab)
        st_fotos_flex = _count_fotos_flex_a_keus(ws_tab)

        ws_stats.cell(row=ri, column=2, value=st_aantal)
        ws_stats.cell(row=ri, column=3, value=st_kar)
        ws_stats.cell(row=ri, column=4, value=st_fotos_dr)
        ws_stats.cell(row=ri, column=5, value=st_fotos_overig)
        ws_stats.cell(row=ri, column=6, value=st_fotos_flex)

        try:
            comp = 25 - (float(st_kar) / 1000.0)
            if st_fotos_dr == 2:
                comp += 2
            if st_fotos_dr == 1:
                comp += 5
            if (st_fotos_dr + st_fotos_overig) < 3.1:
                comp += 1
            if st_aantal < 3.1:
                comp += 1
            if st_fotos_flex == 2:
                comp -= 1
            if st_fotos_flex > 2.9:
                comp -= 2
        except Exception:
            comp = None
        complexity_map[tab] = comp
        ws_stats.cell(row=ri, column=7, value=comp)
        # v28/29: Akpp_ondergrens, Akpp_range, Akpp_range_boost
        akpp_onder = None
        akpp_range = None
        akpp_range_boost = None
        try:
            if st_kar is not None:
                akpp_onder = (float(st_kar) * 0.7) / 2.0
                if akpp_onder > 4500:
                    akpp_onder = 4500
                akpp_onder = int(round(akpp_onder))
                akpp_range = f"{akpp_onder}:7200"
                # v29: boost = min+300 en max-100
                try:
                    _min_s, _max_s = akpp_range.split(":")
                    _min_v = int(float(_min_s))
                    _max_v = int(float(_max_s))
                    akpp_range_boost = f"{_min_v + 300}:{_max_v - 100}"
                except Exception:
                    akpp_range_boost = None
        except Exception:
            akpp_onder = None
            akpp_range = None
            akpp_range_boost = None
        akpp_range_map[tab] = akpp_range
        akpp_range_boost_map[tab] = akpp_range_boost
        ws_stats.cell(row=ri, column=8, value=akpp_onder)
        ws_stats.cell(row=ri, column=9, value=akpp_range)
        ws_stats.cell(row=ri, column=10, value=akpp_range_boost)
        
        
        
            
    # v28/29: kopieer Akpp_range naar AG1 en Akpp_range_boost naar AH1 op elk tabblad
    for tab in ordered_tabs:
        rng = akpp_range_map.get(tab)
        rng_boost = akpp_range_boost_map.get(tab)
        if tab in wb.sheetnames:
            if rng is not None:
                wb[tab]["AG1"].value = rng
            if rng_boost is not None:
                wb[tab]["AH1"].value = rng_boost
    
    # === Tabblad Planningsvolgorde maken (v23) ===
    plan_name = "Planningsvolgorde"
    if plan_name in wb.sheetnames:
        wb.remove(wb[plan_name])
    ws_plan = wb.create_sheet(title=plan_name)

    ws_plan["A1"] = "Planningsvolgorde"

    # Headerkop opmaak (zoals andere tabbladen): Bold + 20% grijs
    _hdr_font = Font(bold=True)
    _hdr_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws_plan["A1"].font = _hdr_font
    ws_plan["A1"].fill = _hdr_fill
    ws_plan["A2"] = "GO-01"
    ws_plan["A5"] = "GO-02"

    other_tabs = ["ZU-MH", "ZU-PS", "ZU-SG", "NM-MI", "NM-NO"]
    other_tabs = [t for t in other_tabs if t in complexity_map]

    other_tabs_sorted = sorted(
        other_tabs,
        key=lambda t: (complexity_map.get(t) is None, -(complexity_map.get(t) or 0)),
    )

    target_rows = [3, 4, 6, 7, 8]
    for r, t in zip(target_rows, other_tabs_sorted):
        ws_plan.cell(row=r, column=1, value=t)

    # === Tabblad Logfile maken (v25) ===
    log_name = "Logfile"
    if log_name in wb.sheetnames:
        wb.remove(wb[log_name])
    ws_log = wb.create_sheet(title=log_name)
    ws_log["A1"] = "Timestamp"
    ws_log["B1"] = "Beschrijving"


    # Headerkop opmaak (zoals andere tabbladen): Bold + 20% grijs
    _hdr_font = Font(bold=True)
    _hdr_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws_log["A1"].font = _hdr_font
    ws_log["A1"].fill = _hdr_fill
    ws_log["B1"].font = _hdr_font
    ws_log["B1"].fill = _hdr_fill
    # === Tabbladen rangschikken (v25) ===
    order_names = []
    for fixed in ["Stats", "Logfile", "Planningsvolgorde", "Totale verhalenlijst"]:
        if fixed in wb.sheetnames:
            order_names.append(fixed)

    # Lees de planningvolgorde (A2 t/m A100) en voeg bestaande tabbladen toe in die volgorde
    planned = []
    for r in range(2, 101):
        v = ws_plan.cell(row=r, column=1).value
        if v is None or str(v).strip() == "":
            continue
        name = str(v).strip()
        if name in wb.sheetnames and name not in planned:
            planned.append(name)

    for name in planned:
        if name not in order_names:
            order_names.append(name)

    # Voeg overige tabbladen toe die nog niet genoemd zijn (volgorde zoals ze nu zijn)
    for ws in wb.worksheets:
        if ws.title not in order_names:
            order_names.append(ws.title)

    # Pas de volgorde toe
    wb._sheets = [wb[n] for n in order_names]



    # v27: output-bestandsnaam = Verhalenaanbod_YYYY-MM-DD.xlsx, datum uit Stats!AA1
    final_output_xlsx = output_xlsx
    try:
        out_dir = Path(output_xlsx).parent if output_xlsx else Path(".")
        aa_val = ws_stats["AA1"].value if "ws_stats" in locals() else None
        if isinstance(aa_val, (datetime.date, datetime.datetime)):
            date_str = aa_val.strftime("%Y-%m-%d")
        elif aa_val is not None and str(aa_val).strip():
            date_str = str(aa_val).strip()[-10:].replace(".", "-")
        else:
            date_str = None
        if date_str:
            final_output_xlsx = str((out_dir / f"Verhalenaanbod_{date_str}.xlsx").resolve())
    except Exception:
        final_output_xlsx = output_xlsx

    wb.save(final_output_xlsx)
    return mapped_rows, warnings, final_output_xlsx

def run_def1(kordiam_report_xlsx: str, mapping_xlsx: str, output_xlsx: str) -> str:
    """Run DEF1 and return the path to the generated Verhalenaanbod/Planningsoverzicht xlsx."""
    mapped_rows, warnings, out_path = process_kordiam(kordiam_report_xlsx, output_xlsx, mapping_xlsx)
    return out_path
