# Auto-generated from DEF3 - PDF Generator.ipynb (cell 5 only)
import os, re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from weasyprint import HTML
from string import Template

def run_def3(*, planning_xlsx: str, mapping_xlsx: str, template_dir: str, out_pdf: str, out_html: str = None) -> str:
    """Run DEF3 and return the path to the generated PDF."""
    PLANNING_XLSX = planning_xlsx
    MAPPING_XLSX = mapping_xlsx
    TEMPLATE_DIR = template_dir
    if out_html is None:
        out_html = os.path.splitext(out_pdf)[0] + '.html'
    OUT_HTML = out_html
    OUT_PDF = out_pdf
    def esc(s):
      if s is None or (isinstance(s, float) and pd.isna(s)): return ""
      return (str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;"))

    def trunc(s, n=30):
      s = "" if s is None or (isinstance(s, float) and pd.isna(s)) else str(s)
      return s if len(s) <= n else s[:n] + "..."

    CODE_RE = re.compile(r"[A-Z]\d{3}[A-Z]")
    def extract_codes(val):
      if val is None or (isinstance(val, float) and pd.isna(val)): return []
      return list(dict.fromkeys(CODE_RE.findall(str(val))))

    def img_path(code):
      p = os.path.join(TEMPLATE_DIR, f"{code}.jpg")
      return p if os.path.exists(p) else None

    wb = load_workbook(PLANNING_XLSX, data_only=True)
    sheet_names = set(wb.sheetnames)

    def get_ae1(sheet):
      if sheet in sheet_names:
        v = wb[sheet]["AE1"].value
        if v is not None and str(v).strip():
          return str(v).strip()
      return sheet

    def get_log_ad(sheet):
      if sheet in sheet_names:
        v = wb[sheet]["AD1"].value
        return "" if v is None else str(v)
      return ""

    # mapping
    m = pd.read_excel(MAPPING_XLSX)
    m.columns = [c.strip() for c in m.columns]
    for c in ["Artikelsoort","Artikel","Beeld"]:
      if c in m.columns:
        m[c] = m[c].astype(str).str.strip()
    map_dict = m.set_index("Artikelsoort")[["Artikel","Beeld"]].to_dict("index")

    # planning print
    df = pd.read_excel(PLANNING_XLSX, sheet_name="Planning print")
    df["Gekozen placeholder"] = df["Gekozen placeholder"].astype(str).str.strip()
    df["Artikel"] = df["Gekozen placeholder"].map(lambda x: map_dict.get(x,{}).get("Artikel",""))
    df["Beeld"]   = df["Gekozen placeholder"].map(lambda x: map_dict.get(x,{}).get("Beeld",""))

    LOG_RULES = [
      ("CM02","<b>Vormgever:</b> Als noodgreep is een stopper-advertentie ingepland van het formaat W32 (104x70 mm). Plaats die eigen uiting of - beter nog - probeer die stopper overbodig te maken door bijvoorbeeld een foto bij een artikel groter te maken."),
      ("CM03","<b>Vormgever:</b> Als noodgreep is een stopper-advertentie ingepland van het formaat W23 (158x94 mm). Plaats die eigen uiting of - beter nog - probeer die stopper overbodig te maken door bijvoorbeeld een foto bij een artikel groter te maken."),
      ("CM04","<b>Vormgever:</b> Als noodgreep is een stopper-advertentie ingepland van het formaat W16 (266x94 mm). Plaats die eigen uiting of - beter nog - probeer die stopper overbodig te maken door bijvoorbeeld een foto bij een artikel groter te maken."),
      ("CI01","<b>Samensteller:</b> Als noodgreep is een artikel S (zonder beeld) open gelaten."),
      ("CH01","<b>Samensteller:</b> Als noodgreep is een artikel XS (zonder beeld) open gelaten."),
      ("CJ01","<b>Samensteller:</b> Als absolute noodgreep is een fors artikel (M over groter) opengelaten. Bekijk wat er de afgelopen dagen is blijven liggen of overleg met de chef hoe deze ruimte kan worden gevuld."),
    ]

    def build_attention_points(plaatsing, g, codes_u):
      points = []
      e_codes = [c for c in codes_u if c.startswith("E")]

      if len(e_codes) == 2:
        points.append("<b>Vormgever:</b> Niet gekozen voor een spreadtemplate, maar voor twee enkele templates. Mogelijk staan beeld en koppen ongelukkig naast elkaar. Doe indien nodig aanpassing.")

      raw_templates = [t for t in g["Gekozen template"].dropna().astype(str).unique() if t.strip()]
      if any("variant" in t.lower() for t in raw_templates):
        points.append("<b>Vormgever en samensteller:</b> Een of meerdere vormen op de template moeten een handmatige reshape krijgen van nieuws naar lichte kop of omgekeerd. Het betreft daarbij per definitie een vorm van maat S of M.")

      log_str = get_log_ad(str(plaatsing))
      for code, text in LOG_RULES:
        if code in log_str:
          points.append(text)

      has_bvp = "Beeld voor print" in g.columns
      for _, r in g.iterrows():
        phs = "" if pd.isna(r.get("Gekozen placeholder")) else str(r.get("Gekozen placeholder"))
        bvp = "" if (not has_bvp) or pd.isna(r.get("Beeld voor print")) else str(r.get("Beeld voor print")).strip()
        tshort = esc(trunc(r.get("Naam productie")))

        if ("B" in phs) and (bvp != "Dragend en bijplaat"):
          points.append(f"<b>Vormgever en samensteller:</b> Bij artikel '{tshort}' is door de chef geen bijplaat gevraagd, maar in de planning kwam het wel goed uit om die toe te kennen. Beoordeel of er inderdaad nog een tweede beeld bij kan. Zo niet, bouw dan de vorm van dit verhaal enigszins om.")

        if (bvp == "Dragend en bijplaat") and ("B" not in phs):
          points.append(f"<b>Vormgever en samensteller:</b> Bij artikel '{tshort}' is door de chef een bijplaat gevraagd naast het dragende beeld, maar deze bijplaat kon bij de planning niet toegekend worden. Beoordeel of dit problematisch is en bouw de vorm van dit verhaal indien nodig enigszins om.")

        if phs.endswith("0") and (bvp not in ["", "Ongeschikt", "Flexibel"]):
          points.append(f"<b>Vormgever en samensteller:</b> Bij artikel '{tshort}' is door de chef ook Beeld gevraagd, maar dit kon bij de planning niet toegekend worden. Grijp alleen in als dit echt problematisch is.")

        if (not phs.endswith("0")) and (bvp == "Ongeschikt"):
          points.append(f"<b>Vormgever en samensteller:</b> Bij artikel '{tshort}' is door de chef aangegeven dat het beeld niet zo geschikt is voor print, maar als absolute noodgreep is er bij de planning toch voor gekozen om bij dit artikel een kleine plaat te gebruiken. In de uitzonderlijke situatie dat het online-beeld bij het verhaal echt niet kan voor print, los je het op met een stopper-advertentie.")

      return points

    def preview_html(codes_u):
      codes = codes_u[:2]
      imgs = [(c, img_path(c)) for c in codes if img_path(c)]
      if not imgs:
        return "<div class='preview'><div class='preview-box'><div style='padding:8px;color:#6b7280;'>Geen preview gevonden voor template.</div></div></div>"

      if len(imgs) == 2 and all(c.startswith('E') for c,_ in imgs):
        return f'''
        <div class="preview">
          <div class="preview-box">
            <table class="preview-table">
              <tr>
                <td><img src="{imgs[0][1]}" alt="{imgs[0][0]}"></td>
                <td class="gap"></td>
                <td><img src="{imgs[1][1]}" alt="{imgs[1][0]}"></td>
              </tr>
            </table>
          </div>
        </div>
        '''
      return f'''
      <div class="preview">
        <div class="preview-box">
          <img class="single" src="{imgs[0][1]}" alt="{imgs[0][0]}">
        </div>
      </div>
      '''


    def fmt_dlabel(v):
      if v is None or (isinstance(v, float) and pd.isna(v)): 
        return ""
      # openpyxl may return datetime/date or a string
      try:
        import datetime as _dt
        if isinstance(v, (_dt.datetime, _dt.date)):
          return _dt.datetime(v.year, v.month, v.day).strftime("%-d %b %Y")
      except Exception:
        pass
      return str(v).strip()

    # Datum voor header komt uit Krantenplanning -> tabblad 'Stats', cel AA1
    date_label = ""
    if "Stats" in sheet_names:
      date_label = fmt_dlabel(wb["Stats"]["AA1"].value)

    # fallback (alleen als Stats/AA1 leeg is)
    if not date_label:
      date_label = " "

    # Genereer-stempel onderaan: datum + tijd van NU
    _now = datetime.now()
    gen_date = _now.strftime("%-d %b %Y")
    gen_time = _now.strftime("%H:%M")
    footer_text = f"Planning gegenereerd met Krantenplanner V1.1, op {gen_date} om {gen_time}"

    cards = []
    for plaatsing, g in df.groupby("Plaatsing", sort=True):
      codes=[]
      for t in g["Gekozen template"]:
        codes += extract_codes(t)
      codes_u = list(dict.fromkeys(codes))

      e_codes = [c for c in codes_u if c.startswith("E")]
      is_spread = (any(c.startswith("S") for c in codes_u) or len(e_codes) == 2)
      typ = "SPREAD" if is_spread else "ENKELE PAGINA"
      tpl = " / ".join(codes_u[:2]) + (" …" if len(codes_u) > 2 else "") if codes_u else "—"

      title = esc(get_ae1(str(plaatsing))).upper()
      meta  = f"{esc(plaatsing)} • {typ} • TEMPLATE {esc(tpl)} • {len(g)} ARTIKELEN"

      rows = ""
      for _, r in g.iterrows():
        rows += f'''
        <tr>
          <td class="col-small">{esc(r.get('Artikel'))}</td>
          <td class="col-medium">{esc(r.get('Beeld'))}</td>
          <td>{esc(r.get('Naam productie'))}</td>
          <td class="col-author">{esc(r.get('Auteur'))}</td>
          <td class="col-region">{esc(r.get('Focusregio'))}</td>
        </tr>
        '''

      prev = preview_html(codes_u)
      pts  = build_attention_points(plaatsing, g, codes_u)
      att  = ""
      if pts:
        att = "<div class='attention'><div class='attention-title'>AANDACHTSPUNTEN</div><ul>" + "".join([f"<li>{p}</li>" for p in pts]) + "</ul></div>"

      bottom = f"<div class='bottom'>{prev}{att}</div>" if (prev or att) else ""

      cards.append(f'''
      <div class="card">
        <div class="title">{title}</div>
        <div class="meta">{meta}</div>

        <table class="data">
          <thead>
            <tr>
              <th class="col-small">Artikel</th>
              <th class="col-medium">Beeld</th>
              <th>Titel</th>
              <th class="col-author">Auteur</th>
              <th class="col-region">Focusregio</th>
            </tr>
          </thead>
          <tbody>{rows}</tbody>
        </table>

        {bottom}
      </div>
      ''')

    cards_html = "<div class='pagebreak'></div>".join(cards)

    def placeholder_to_beeld(ph):
      v = "" if ph is None or (isinstance(ph, float) and pd.isna(ph)) else str(ph).strip()
      m = re.search(r"(\d)\s*$", v)
      if not m:
        return ""
      d = int(m.group(1))
      return "Geen" if d == 0 else f"{d} kolom"

    def placeholder_to_artikel(ph):
      v = "" if ph is None or (isinstance(ph, float) and pd.isna(ph)) else str(ph).strip()
      if not v:
        return ""
      parts = v.split("_")
      # drop trailing numeric part like "_4"
      if parts and re.fullmatch(r"\d+", parts[-1] or ""):
        parts = parts[:-1]
      if not parts:
        return ""
      base = parts[0]
      if len(parts) > 1:
        toevoeging = "_".join(parts[1:])
        return f"{base} ({toevoeging})"
      return base

    def _pick_nonempty(existing, fallback):
      ex = "" if existing is None or (isinstance(existing, float) and pd.isna(existing)) else str(existing).strip()
      return ex if ex else fallback

    def verv_section(sheet_name, heading):
      if sheet_name not in sheet_names:
        return ""
      dfv = pd.read_excel(PLANNING_XLSX, sheet_name=sheet_name).fillna("")
      # ensure expected columns exist
      for col in ["Plaatsing","Gekozen placeholder","Artikel","Beeld","Naam productie","Auteur","Focusregio"]:
        if col not in dfv.columns:
          dfv[col] = ""
      # Fill Artikel/Beeld from Gekozen placeholder when blank
      dfv = dfv.copy()
      dfv["Artikel"] = dfv.apply(lambda r: _pick_nonempty(r.get("Artikel",""), placeholder_to_artikel(r.get("Gekozen placeholder",""))), axis=1)
      dfv["Beeld"]   = dfv.apply(lambda r: _pick_nonempty(r.get("Beeld",""),   placeholder_to_beeld(r.get("Gekozen placeholder",""))), axis=1)

      cols = [("Plaatsing","Positie"),("Artikel","Artikel"),("Beeld","Beeld"),("Naam productie","Titel"),("Auteur","Auteur"),("Focusregio","Regio")]
      header = "".join([f"<th>{dst}</th>" for _,dst in cols])
      rows = ""
      for _, r in dfv.iterrows():
        rows += "<tr>" + "".join([f"<td>{esc(r.get(src,''))}</td>" for src,_ in cols]) + "</tr>"
      return f'''
        <div style="margin-top:18px;">
          <div style="font-weight:800;margin:0 0 8px 0;">{heading}</div>
          <table class="data">
            <thead><tr>{header}</tr></thead>
            <tbody>{rows}</tbody>
          </table>
        </div>
      '''

    appendix_html = f'''
    <div class="pagebreak"></div>
    <div class="card">
      <div class="title">BIJLAGE 1: WAT TE VERVANGEN BIJ LAAT NIEUWS?</div>
      <div class="meta">Indien er in de loop van de avond extra verhalen bij komen ten opzichte van de originele planning, dan kunnen de volgende verhalen wijken:</div>
      {verv_section("NM-VERV","In Noord-Midden")}
      {verv_section("ZU-VERV","In Zuid")}
      {verv_section("GO-VERV","Op gehele oplage")}
    </div>
    '''



    def artikel_label(val):
      v = "" if val is None or (isinstance(val, float) and pd.isna(val)) else str(val).strip()
      if "_" in v:
        a, b = v.split("_", 1)
        return f"{a} ({b})"
      return v

    def unused_section(sheet_name, heading):
      if sheet_name not in sheet_names:
        return ""
      dfu = pd.read_excel(PLANNING_XLSX, sheet_name=sheet_name).fillna("")
      # normalize expected columns
      for col in ["Artikelsoort","Beeld voor print","Naam productie","Auteur","Focusregio"]:
        if col not in dfu.columns:
          dfu[col] = ""
      dfu = dfu.copy()
      dfu["Artikel"] = dfu["Artikelsoort"].map(artikel_label)
      cols = [("Artikel","Artikel"), ("Beeld voor print","Beeld"), ("Naam productie","Titel"), ("Auteur","Auteur"), ("Focusregio","Focusregio")]
      header = "".join([f"<th>{esc(dst)}</th>" for _,dst in cols])
      rows = ""
      for _, r in dfu.iterrows():
        rows += "<tr>" + "".join([f"<td>{esc(r.get(src,''))}</td>" for src,_ in cols]) + "</tr>"
      return f'''
        <div style="margin-top:18px;">
          <div style="font-weight:800;margin:0 0 8px 0;">{heading}</div>
          <table class="data">
            <thead><tr>{header}</tr></thead>
            <tbody>{rows}</tbody>
          </table>
        </div>
      '''

    appendix2_html = f'''
    <div class="pagebreak"></div>
    <div class="card">
      <div class="title">BIJLAGE 2: DEZE VERHALEN VIELEN OVER DE RAND</div>
      <div class="meta">Deze verhalen waren wel aangemerkt als bruikbaar, maar hebben bij de planning geen plek toegebeeld gekregen. Dat kan zijn omdat ze weinig prioriteit hadden, maar het kan ook zijn dat dit verhaal steeds pech had dat per pagina met dit verhaal steeds de 'puzzel' niet paste. Beoordeel of er op deze lijst onverhoopt verhalen staan die alsnog mee moeten, en grijp indien nodig in.</div>
      {unused_section("NM-UNUSED","In Noord-Midden:")}
      {unused_section("ZU-UNUSED","In Zuid:")}
    </div>
    '''

    tpl = Template(r'''
    <!doctype html>
    <html>
    <head>
    <meta charset="utf-8">
    <style>
      @page { size: A4 landscape; margin: 18mm 15mm 18mm 15mm;
        @top-left { content: "Krantenplanning"; font-size: 10pt; font-weight: 700; }
        @top-center { content: "De Limburger · $date_label"; font-size: 10pt; font-weight: 600; color: #333; }
        @top-right { content: "Pagina " counter(page) " / " counter(pages); font-size: 10pt; color: #333; }
          @bottom-center { content: "$footer_text"; font-size: 10pt; font-weight: 600; color: #333; }
    }
      body { font-family: Arial, sans-serif; color:#111; font-size: 11px; }

      .content { margin-top: 0; }

      .card {
        background:#f4f8fb; border: 1px solid #e5eef6; border-radius: 10px;
        padding: 12px 14px;
      }
      .title { font-size: 20px; font-weight: 800; margin: 0; }
      .meta { margin-top: 6px; color:#4b5563; font-weight: 600; }

      table.data { width:100%; border-collapse: collapse; margin-top: 10px; background:#fff; border-radius: 8px; overflow:hidden; }
      table.data thead th {
        text-align:left; background:#e6f0f7; color:#0f172a;
        padding: 7px 8px; font-size: 11px; border-bottom: 1px solid #cfe0ee;
      }
      table.data tbody td { padding: 7px 8px; border-bottom: 1px solid #eef2f6; vertical-align: top; }
      table.data tbody tr:last-child td { border-bottom: none; }

      .col-small { width: 55px; white-space: nowrap; }
      .col-medium { width: 90px; white-space: nowrap; }
      .col-author { width: 160px; }
      .col-region { width: 95px; white-space: nowrap; }

      .bottom { display:flex; gap: 14px; margin-top: 12px; align-items:flex-start; }

      .preview {
        border:1px solid #d1d5db; border-radius: 6px;
        padding: 6px; background: #fff; width: 310px;
        overflow: hidden; box-sizing: border-box;
      }
      .preview img { display:block; border-radius: 4px; }
      .preview-box { max-height: 78mm; overflow: hidden; }
      .preview-table { width:100%; border-collapse: collapse; table-layout: fixed; }
      .preview-table td { padding:0; vertical-align: top; text-align: center; }
      .preview-table td.gap { width:2mm; }
      .preview-table img { max-height: 78mm; max-width: 100%; width: auto; height: auto; object-fit: contain; display:inline-block; border-radius: 6px; }
      .preview-box img.single { max-height: 78mm; max-width: 100%; width: auto; height: auto; object-fit: contain; display:block; margin: 0 auto; border-radius: 6px; }

      .attention {
        background:#fff; border: 1px solid #e5e7eb; border-radius: 10px;
        padding: 10px 12px; flex: 1;
      }
      .attention-title { font-weight: 800; margin-bottom: 6px; }
      .attention ul { margin: 0; padding-left: 18px; }
      .attention li { margin-bottom: 6px; }

      .pagebreak { page-break-after: always; }
      .pagebreak:last-child { page-break-after: auto; }
    </style>
    </head>
    <body>

      <div class="content">
        $cards_html
        $appendix_html
      </div>
    </body>
    </html>
    ''')

    html = tpl.substitute(date_label=esc(date_label), footer_text=esc(footer_text), cards_html=cards_html, appendix_html=appendix_html + appendix2_html)

    with open(OUT_HTML, "w", encoding="utf-8") as f:
      f.write(html)

    HTML(string=html, base_url=os.getcwd()).write_pdf(OUT_PDF)
    print("✅ PDF gemaakt:", OUT_PDF)
    return OUT_PDF
