# Auto-generated from DEF2 - Krantenplanner.ipynb
import openpyxl
import pandas as pd
import re
import itertools
import datetime
import os

def run_def2(*, templates_path: str, beslispad_spread_path: str, beslispad_ep_path: str, posities_path: str, verhalenaanbod_path: str, out_path: str) -> str:
    """Run DEF2 and return the path to the generated Krantenplanning xlsx."""
    TEMPLATES_PATH = templates_path
    BESLISPAD_SPREAD_PATH = beslispad_spread_path
    BESLISPAD_EP_PATH = beslispad_ep_path
    POSITIES_PATH = posities_path
    VERHALENAANBOD_PATH = verhalenaanbod_path
    def copy_storylist(wb, src_name, dst_name):
        if src_name not in wb.sheetnames:
            return
        if dst_name in wb.sheetnames:
            wb.remove(wb[dst_name])
        wb.copy_worksheet(wb[src_name]).title = dst_name

    def overwrite_row_by_name(ws, name, vals):
        # Overschrijf bestaande rij (op Naam productie), behoud kolomstructuur
        headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
        name_col = None
        for i,h in enumerate(headers, start=1):
            if str(h).strip().lower() == "naam productie":
                name_col = i
                break
        if name_col is None:
            return
        for r in range(2, ws.max_row+1):
            if normalize(ws.cell(r, name_col).value) == normalize(name):
                for c,h in enumerate(headers, start=1):
                    if h in vals:
                        ws.cell(r,c).value = vals[h]
                return
    #@title RUN – Genereer Krantenplanning.xlsx
    # Deze cel voert de matching-engine uit en schrijft Krantenplanning.xlsx weg.
    # Daarna wordt het bestand automatisch aangeboden voor download.


    def _header_index(ws):
        hdr = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
        norm = [normalize(h) for h in hdr]
        return hdr, {n:i+1 for i,n in enumerate(norm)}

    def _get_col(idx_map, *names):
        for nm in names:
            key = normalize(nm)
            if key in idx_map:
                return idx_map[key]
        return None

    def normalize(x):
        return "" if x is None else str(x).strip()

    def truthy(v):
        if v is None: return False
        if isinstance(v, bool): return v
        if isinstance(v, (int, float)): return v != 0
        return str(v).strip().lower() in ("true","ja","1","yes","waar")

    def cell_set(v):
        # Splits cellen zoals "S_nws_0; S_nws_2;" naar set({"S_nws_0","S_nws_2"})
        if v is None: return set()
        return set([p.strip() for p in str(v).split(";") if p.strip()])

    def find_col_any(df, substrings, required=True):
        for sub in substrings:
            for c in df.columns:
                if sub.lower() in str(c).lower():
                    return c
        raise KeyError(f"Geen kolom gevonden voor: {substrings}")

    def _pretty_class_token(t):
        t = normalize(t)
        if t == "a-keus":
            return "A-keus"
        if t == "b-keus":
            return "B-keus"
        return t


    def _split_multi(val):
        if val is None:
            return set()
        s = str(val).strip()
        if s == "":
            return set()
        # accepteer ; , + / en ' en ' als scheiding
        parts = re.split(r"[;,+/]|en", s, flags=re.IGNORECASE)
        return {normalize(p) for p in parts if normalize(p)}

    def class_allowed(pc, allowed):
        # allowed kan zijn "Alle" of bv. "A-keus" of meerdere waarden
        if allowed is None or str(allowed).strip()=="":
            return True
        if str(allowed).strip().lower()=="alle":
            return True
        allowed_set = _split_multi(allowed)
        pc_set = _split_multi(pc)
        # match zodra er overlap is
        return len(pc_set & allowed_set) > 0


    def parse_range(s):
        if s is None: return None
        s=str(s).strip()
        m=re.match(r"^\s*(\d+(?:\.\d+)?)\s*:\s*(\d+(?:\.\d+)?)\s*$", s)
        if not m: return None
        return (float(m.group(1)), float(m.group(2)))

    # -----------------------------
    # Load workbooks (data_only=True is essentieel i.v.m. formules zoals Posities!M1)
    # -----------------------------
    plan_wb = openpyxl.load_workbook(VERHALENAANBOD_PATH, data_only=True)
    pos_wb  = openpyxl.load_workbook(POSITIES_PATH, data_only=True)
    tmpl_wb = openpyxl.load_workbook(TEMPLATES_PATH, data_only=True)

    pos_ws   = pos_wb["Blad1"]
    # Helper: always resolve the current Logfile worksheet (important after fallback reset)
    def get_log_ws():
        if "Logfile" not in plan_wb.sheetnames:
            plan_wb.create_sheet("Logfile", 0)
            ws = plan_wb["Logfile"]
            ws.append(["Timestamp","Beschrijving"])
        return plan_wb["Logfile"]

    # Reset logfile (laat header staan)
    _log_ws = get_log_ws()
    if _log_ws.max_row > 1:
        _log_ws.delete_rows(2, _log_ws.max_row)

    def log(msg):
        ws = get_log_ws()
        row = ws.max_row + 1
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=row, column=1, value=ts)
        ws.cell(row=row, column=2, value=msg)

    # -----------------------------
    # Planning order
    # -----------------------------
    po_ws = plan_wb["Planningsvolgorde"]
    order = []
    r = 2
    while True:
        v = po_ws[f"A{r}"].value
        if v is None or str(v).strip()=="":
            break
        order.append(str(v).strip())
        r += 1

    # -----------------------------
    # -----------------------------
    # Regime berekenen (VERWIJDERD)
    # In deze versie is er geen onderscheid meer tussen Normaal/Verhalenschaarste/Papierschaarste.
    # Akpp-range wordt per run afgelezen vanaf het run-tabblad (AG1/AH1).
    # -----------------------------

    # Templates inlezen
    # -----------------------------
    tmpl_ws = tmpl_wb["Blad1"]
    tmpl_headers = [tmpl_ws.cell(1,c).value for c in range(1, tmpl_ws.max_column+1)]
    def tcol(name): return tmpl_headers.index(name)+1

    templates = []
    for rr in range(2, tmpl_ws.max_row+1):
        tpl = tmpl_ws.cell(rr, tcol("Template")).value
        if tpl is None: 
            continue
        placeholders = []
        for i in range(1,6):
            v = tmpl_ws.cell(rr, tcol(f"Placeholder {i}")).value
            if v not in (None,""):
                placeholders.append(str(v).strip())
        templates.append({
            "Template": str(tpl).strip(),
            "Templatesoort": normalize(tmpl_ws.cell(rr, tcol("Templatesoort")).value),
            "Placeholders": placeholders,
            "Advertentiepositie": normalize(tmpl_ws.cell(rr, tcol("Advertentiepositie")).value),
            "Akpp": tmpl_ws.cell(rr, tcol("Akpp")).value,
        })

    # -----------------------------
    # Beslispaden inlezen (pandas)
    # -----------------------------
    bps = pd.read_excel(BESLISPAD_SPREAD_PATH, sheet_name="Blad1")
    bpe = pd.read_excel(BESLISPAD_EP_PATH, sheet_name="Blad1")

    # Zorg dat stapcodes als tekst blijven (BPS1.001 / BPE1.001)
    if 'Stappen' in bps.columns:
        bps['Stappen'] = bps['Stappen'].astype(str)
    if 'Stappen' in bpe.columns:
        bpe['Stappen'] = bpe['Stappen'].astype(str)

    def map_cols(df):
        # Robust tegen aangepaste kolomkoppen (zoals jouw update)
        return {
            "mode": find_col_any(df, ["Bovenste producties"]),
            "skip_pos": find_col_any(df, ["Sla deze stap over bij deze posities"], required=False),
            "class": find_col_any(df, ["Toegestane Classificatie"]),
            "tsoort": find_col_any(df, ["Toegestane 'Templatesoort'", "Templatesoort"]),
            # Spread-kop kan door updates veranderen, daarom zoeken we breed:
            "tw1_if": find_col_any(df, ["Beeld voor print=Bijplaat", "wordt voldaan", "toegestaan if"]),
            "tw1": find_col_any(df, ["Bij maximaal 1 van de placeholders op het template 'Tweede keus placeholder' toegestaan "]),
            "tw2": find_col_any(df, ["Bij maximaal 2 van de placeholders"]),
            "open_xs": find_col_any(df, ["XS_0 open"]),
            "open_s": find_col_any(df, ["S_nws_0 of S_lk_0 open", "S_nws_0"]),
            "open_custom": find_col_any(df, ["Toegestaan om maximaal 1 placeholder op te laten van onderstaande soort(en)"]),
            "derde": find_col_any(df, ["Derde keus placeholder"]),
            "derde_max1": find_col_any(df, ["Bij maximaal 1 van de placeholders op het template \'Derde keus placeholder\' toegestaan"]),
            "vierde": find_col_any(df, ["Bij maximaal 1 van de placeholders op het template 'VIERDE keus placeholder' toegestaan"]),
            "admatch": find_col_any(df, ["Advertentiepositie' matchen", "Advertentiepositie"]),
            "akpp": find_col_any(df, ["Uitsluitend templates toegestaan met Akpp-waarde binnen deze range", "Akpp-waarde binnen deze range", "Akpp range"], required=False),
            "conct": find_col_any(df, ["Concessies_beschreven"]),
        }

    cols_bps = map_cols(bps)
    cols_bpe = map_cols(bpe)

    def resolve_akpp_range(code_or_range, tabname):
        """Los een Akpp-range op voor deze run.

        - Directe range: 'min:max' -> (min,max)
        - Code 'Akpp_range' -> lees van tabblad <tabname> cel AG1
        - Code 'Akpp_range_boost' -> lees van tabblad <tabname> cel AH1
        """
        if code_or_range is None:
            return None
        s = str(code_or_range).strip()
        # 1) direct 'min:max'
        direct = parse_range(s)
        if direct:
            return direct

        code = s.lower().replace(" ", "")
        # 2) via run-tabblad
        try:
            ws_run = plan_wb[tabname]
        except Exception:
            ws_run = None

        if code in ("akpp_range", "akpprange"):
            if ws_run is None:
                return None
            return parse_range(ws_run["AG1"].value)

        if code in ("akpp_range_boost", "akpprange_boost", "akpprangeboost"):
            if ws_run is None:
                return None
            return parse_range(ws_run["AH1"].value)

        # Onbekende code: geen Akpp-filter
        return None

    # -----------------------------
    # Posities kolommen
    # -----------------------------
    pos_header = [pos_ws.cell(1,c).value for c in range(1, pos_ws.max_column+1)]
    def pos_col(sub):
        for i,h in enumerate(pos_header, start=1):
            if h and sub.lower() in str(h).lower():
                return i
        raise KeyError(sub)

    POS_VORM_COL = pos_col("Verschijningsvorm")
    POS_POS_COL  = pos_col("Positie")
    POS_AD1_COL  = pos_col("Advertentieaanbod")
    POS_AD2_COL  = pos_col("tweede keus")
    POS_AD3_COL  = pos_col("derde keus")
    POS_AD4_COL  = pos_col("vierde keus")

    def get_pos_row(posname):
        rr=2
        while True:
            v = pos_ws.cell(rr, POS_POS_COL).value
            if v is None or str(v).strip()=="":
                return None
            if str(v).strip()==posname:
                return rr
            rr += 1

    # -----------------------------
    # Sheet <-> DataFrame
    # -----------------------------
    def sheet_to_df(ws):
        headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
        data=[]
        for rr in range(2, ws.max_row+1):
            row = [ws.cell(rr,c).value for c in range(1, ws.max_column+1)]
            if all(v is None or str(v).strip()=="" for v in row):
                continue
            data.append(row)
        return pd.DataFrame(data, columns=headers)

    def df_to_sheet(ws, df):
        ws.delete_rows(2, ws.max_row)
        for i,row in enumerate(df.itertuples(index=False), start=2):
            for j,val in enumerate(row, start=1):
                ws.cell(i,j,value=val)

    def get_copy_targets(ws):
        # Tabbladen om NAAR TE KOPIËREN (stap 7) – uit cel AA1
        v = ws["AA1"].value
        return [s.strip() for s in re.split(r"[;,]", str(v)) if s.strip()] if v else []

    def get_keep_targets(ws):
        # Tabbladen die NIET gestript mogen worden (stap 8) – uit cel AB1
        v = ws["AB1"].value
        return [s.strip() for s in re.split(r"[;,]", str(v)) if s.strip()] if v else []

        rr = ws.max_row + 1
        for c,val in enumerate(vals, start=1):
            ws.cell(rr,c,value=val)

    def remove_from_sheet(ws, name):
        # Verwijder rij op basis van kolom 'Naam productie' (ongeacht kolomvolgorde)
        headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
        name_col = None
        for i,h in enumerate(headers, start=1):
            if normalize(h).lower() == "naam productie":
                name_col = i
                break
        if name_col is None:
            name_col = 1  # fallback: oude structuur

        target = normalize(name)
        for rr in range(2, ws.max_row+1):
            if normalize(ws.cell(rr, name_col).value) == target:
                ws.delete_rows(rr,1)
                return

    # -----------------------------
    # Matching per stap
    # -----------------------------
    def best_match_for_step(tabname, df_run, step_row, is_spread):
        cols = cols_bps if is_spread else cols_bpe

        mode          = step_row[cols["mode"]]
        is_bovenste   = str(mode).lower().startswith("bovenste")
        allowed_class = step_row[cols["class"]]
        allowed_tsoort= step_row[cols["tsoort"]]
        tw1_if        = truthy(step_row[cols["tw1_if"]])
        tw1           = truthy(step_row[cols["tw1"]])
        tw2           = truthy(step_row[cols["tw2"]])
        open_xs       = truthy(step_row[cols["open_xs"]])
        open_s        = truthy(step_row[cols["open_s"]])
        open_custom_raw = step_row[cols["open_custom"]] if "open_custom" in cols else None
        open_custom_set = cell_set(open_custom_raw)
        derde         = truthy(step_row[cols["derde"]])
        derde_max1    = truthy(step_row[cols["derde_max1"]])
        vierde        = truthy(step_row[cols["vierde"]])
        admatch       = normalize(step_row[cols["admatch"]])
        conc_txt      = step_row[cols["conct"]]

        rng_cell = step_row[cols.get("akpp")] if cols.get("akpp") is not None else None
        rng_pair = resolve_akpp_range(rng_cell, tabname)

        pr = get_pos_row(tabname)
        ad1 = normalize(pos_ws.cell(pr, POS_AD1_COL).value) if pr else ""
        ad2 = normalize(pos_ws.cell(pr, POS_AD2_COL).value) if pr else ""
        ad3 = normalize(pos_ws.cell(pr, POS_AD3_COL).value) if pr else ""
        ad4 = normalize(pos_ws.cell(pr, POS_AD4_COL).value) if pr else ""
        ad_choice = {"Advertentieaanbod":ad1, "Advertentieaanbod_tweede keus":ad2, "Advertentieaanbod_derde keus":ad3, "Advertentieaanbod_vierde keus":ad4}.get(admatch, ad1)

        # IF Advertentieaanbod != W00 THEN range-min -1000
        if rng_pair and ad_choice and ad_choice!="W00":
            rng_pair = (rng_pair[0]-1000, rng_pair[1])

        def template_allowed(t):
            if allowed_tsoort:
                # Sta ";" en "," toe als scheiding in Toegestane Templatesoort
                allowed_tsoort_set = {normalize(p) for p in re.split(r"[;,]", str(allowed_tsoort)) if str(p).strip()}
                if normalize(t["Templatesoort"]) not in allowed_tsoort_set:
                    return False
            if ad_choice and normalize(t["Advertentiepositie"]) != ad_choice:
                return False
            if rng_pair:
                try:
                    ak = float(t["Akpp"])
                except:
                    return False
                return rng_pair[0] <= ak <= rng_pair[1]
            return True

        tpls = [t for t in templates if template_allowed(t)]
        if not tpls:
            return None

        pool = df_run.copy()
        if "Classificatie" in pool.columns:
            pool = pool[pool["Classificatie"].apply(lambda x: class_allowed(x, allowed_class))]
        pool = pool.reset_index(drop=True)
        if pool.empty:
            return None

        max_tw  = 2 if tw2 else (1 if (tw1 or tw1_if) else 0)
        max_der = 0 if not derde else (1 if derde_max1 else len(phs))
        max_v   = 1 if vierde else 0

        # Precompute placeholder sets
        for col in ["Gewenste placeholder","Tweede keus placeholder","Derde keus placeholder","Vierde keus placeholder"]:
            pool[col+"_set"] = pool[col].apply(cell_set) if col in pool.columns else [set()]*len(pool)

        def matches(prow, ph, kind):
            return ph in prow[{"g":"Gewenste placeholder_set","t":"Tweede keus placeholder_set","d":"Derde keus placeholder_set","v":"Vierde keus placeholder_set"}[kind]]

        best = None
        for t in tpls:
            phs = t["Placeholders"]
            k = len(phs)
            if is_bovenste:
                # Neem bovenste N producties (N = aantal placeholders),
                # maar bij ex aequo Prioscore: neem ook alle extra producties met dezelfde Prioscore
                pool_use = pool
                if k < len(pool_use):
                    try:
                        boundary = float(pool_use.loc[k-1].get("Prioscore", None))
                        last = k-1
                        while last + 1 < len(pool_use):
                            nxt = pool_use.loc[last+1].get("Prioscore", None)
                            try:
                                nxt_f = float(nxt)
                            except Exception:
                                break
                            if nxt_f == boundary:
                                last += 1
                            else:
                                break
                        pool_use = pool_use.iloc[:last+1]
                    except Exception:
                        pool_use = pool_use.iloc[:k]
            else:
                pool_use = pool

            def _bovenste_ok(selected_indices):
                # Regels voor 'Bovenste' met ex aequo:
                # - als je r producties gebruikt, dan mag je NIET een hogere prioscore overslaan.
                # - je mag alleen wisselen binnen de boundary-tie (zelfde Prioscore als de r-de productie).
                try:
                    r = len(selected_indices)
                    if r == 0:
                        return True
                    boundary = float(pool.loc[r-1].get("Prioscore", None))
                    # verplicht: alle indices met Prioscore > boundary moeten geselecteerd zijn (prefix)
                    mandatory_count = 0
                    for j in range(len(pool)):
                        try:
                            pj = float(pool.loc[j].get("Prioscore", None))
                        except Exception:
                            break
                        if pj > boundary:
                            mandatory_count += 1
                        else:
                            break
                    for j in range(mandatory_count):
                        if j not in selected_indices:
                            return False
                    # toegestaan: indices t/m last waarbij Prioscore == boundary (ties)
                    last = r-1
                    while last + 1 < len(pool):
                        try:
                            nxt = float(pool.loc[last+1].get("Prioscore", None))
                        except Exception:
                            break
                        if nxt == boundary:
                            last += 1
                        else:
                            break
                    return all((i <= last) for i in selected_indices)
                except Exception:
                    # fallback: strict prefix-regel zonder ties
                    r = len(selected_indices)
                    if r == 0:
                        return True
                    return set(range(r)).issubset(set(selected_indices))

            # Candidate list per placeholder
            candidates=[]
            for ph in phs:
                c=[]
                for idx, prow in pool_use.iterrows():
                    if matches(prow, ph, "g"):
                        c.append((idx,"g")); 
                        continue
                    if max_tw>0:
                        if tw1_if:
                            # Let op: dit blijft exact "bijplaat" (zoals in de huidige output-logica)
                            if normalize(prow.get("Beeld voor print","")).lower()=="bijplaat" and matches(prow, ph, "t"):
                                c.append((idx,"t"))
                        else:
                            if matches(prow, ph, "t"):
                                c.append((idx,"t"))
                    if max_der>0 and matches(prow, ph, "d"):
                        c.append((idx,"d"))
                    if max_v>0 and matches(prow, ph, "v"):
                        c.append((idx,"v"))
                if ph=="XS_0" and open_xs:
                    c.append((None,"o"))
                if ph in ("S_nws_0","S_lk_0") and open_s:
                    c.append((None,"o"))
                if open_custom_set and ph in open_custom_set:
                    c.append((None,"o"))
                candidates.append(c)

            if any(len(c)==0 for c in candidates):
                continue

            # Brute force (max 5 placeholders)
            best_for_tpl=None
            for choice in itertools.product(*candidates):
                idxs=[i for i,_k in choice if i is not None]
                if len(idxs)!=len(set(idxs)): 
                    continue  # productie mag niet dubbel
                if sum(1 for _i,k in choice if k=="t")>max_tw: 
                    continue
                if sum(1 for _i,k in choice if k=="d")>max_der:
                    continue
                if sum(1 for _i,k in choice if k=="v")>max_v:
                    continue

                # MAXIMAAL 1× S_nws_0 of S_lk_0 open laten (samen)
                open_s_count = sum(
                    1 for slot,(i,k) in enumerate(choice)
                    if i is None and k=="o" and phs[slot] in ("S_nws_0","S_lk_0")
                )
                if open_s_count > 1:
                    continue

                # MAXIMAAL 1× open laten uit custom lijst (uit beslispad)
                if open_custom_set:
                    custom_open_count = sum(
                        1 for slot,(i,k) in enumerate(choice)
                        if i is None and k=="o" and phs[slot] in open_custom_set
                    )
                    if custom_open_count > 1:
                        continue

                prios=[]
                for i,_k in choice:
                    if i is None: 
                        continue
                    try:
                        prios.append(float(pool_use.loc[i].get("Prioscore",0)))
                    except:
                        prios.append(0.0)
                if not prios:
                    continue

                metric=(sum(prios)/len(prios), sum(prios))  # avg, sum
                if best_for_tpl is None or metric>best_for_tpl["metric"]:
                    best_for_tpl={"metric":metric, "choice":choice, "pool_use":pool_use}

            if best_for_tpl and (best is None or best_for_tpl["metric"]>best["metric"]):
                best={"metric":best_for_tpl["metric"], "template":t, "choice":best_for_tpl["choice"], "pool_use":best_for_tpl["pool_use"], "conct":conc_txt}

        return best

    # -----------------------------
    # Execute pipeline
    # -----------------------------
    success = 0


    def _apply_akpp_range_berekening(ws_run, df_run, tabname):
        """
        [AKPP_RANGE_BEREKENING] (versie 16)
        - Karakters_beschikbaar = som(Karakters) voor alle rijen, maar tel NIET mee als 'Telt mee' = nee
        - Restant = [POSITIELIJST] kolom 'Restant' voor dit tabblad
        - Karakters_gewenst_pp = Karakters_beschikbaar / Restant
        - Ondergrens = clip(Karakters_gewenst_pp - 1250, 2700..5200)
        - Bovengrens = clip(Karakters_gewenst_pp + 1250, 5300..7800)
        - Akpp_range = "min:max" -> schrijf naar AG1
        - Akpp_range_boost: min+400 : max-400 -> schrijf naar AH1
        """
        # vind kolommen in df (header-based)
        c_kar = _find_df_col(df_run, "Karakters")
        c_tm  = _find_df_col(df_run, "Telt mee")  # kan ontbreken
        if c_kar is None:
            return

        def _to_float(x):
            try:
                if x is None or (isinstance(x, float) and (x != x)):
                    return 0.0
                s = str(x).strip().replace(",", ".")
                return float(s)
            except Exception:
                return 0.0

        kar_sum = 0.0
        for _, row in df_run.iterrows():
            if c_tm is not None:
                tm = normalize(row.get(c_tm, "")).strip().lower()
                if tm == "nee":
                    continue
            kar_sum += _to_float(row.get(c_kar, 0))

        # Restant uit POSITIELIJST
        rest_col = None
        for i, h in enumerate(pos_header, start=1):
            if normalize(h).lower() == "restant":
                rest_col = i
                break
        if rest_col is None:
            return
        pr = get_pos_row(tabname)
        if pr is None:
            return
        try:
            restant = int(float(pos_ws.cell(pr, rest_col).value))
        except Exception:
            return
        if restant <= 0:
            return

        gewenst_pp = kar_sum / restant

        onder = max(2700.0, min(5200.0, gewenst_pp - 1250.0))
        boven = max(5300.0, min(7800.0, gewenst_pp + 1250.0))

        onder_i = int(round(onder))
        boven_i = int(round(boven))

        ws_run["AG1"].value = f"{onder_i}:{boven_i}"

        boost_min = onder_i + 400
        boost_max = boven_i - 400
        ws_run["AH1"].value = f"{boost_min}:{boost_max}"


    def run_runs(runs):
        global success
        for posname in runs:
            if posname not in plan_wb.sheetnames:
                log(f"Run {posname}: tabblad ontbreekt.")
                continue

            pr = get_pos_row(posname)
            vorm = normalize(pos_ws.cell(pr, POS_VORM_COL).value) if pr else ""

            if vorm.lower()=="niet":
                plan_wb.remove(plan_wb[posname])
                log(f"Run {posname}: Verschijningsvorm=Niet -> verwijderd.")
                continue

            is_spread = (vorm.lower()=="spread")
            beslis = bps if is_spread else bpe
            cols_step = cols_bps if is_spread else cols_bpe

            ws = plan_wb[posname]
            df = sheet_to_df(ws)

            # Speciaal voor NM-U1: bepaal Akpp_range/Akpp_range_boost volgens [AKPP_RANGE_BEREKENING]
            if posname in ("NM-U1","NM-U2","NM-U3","NM-U4","NM-U5","NM-U6","ZU-U1","ZU-U2","ZU-U3","ZU-U4","ZU-U5","ZU-U6"):
                _apply_akpp_range_berekening(ws, df, posname)
            copy_targets = get_copy_targets(ws)
            keep_targets = get_keep_targets(ws)

            matched=False
            for _, step in beslis.iterrows():
                # Optioneel: sla deze stap over voor specifieke posities
                if "skip_pos" in cols_step and cols_step["skip_pos"] is not None:
                    raw = step[cols_step["skip_pos"]]
                    skip_set = {normalize(s) for s in re.split(r"[;,]", str(raw)) if str(s).strip()} if raw not in (None, "") else set()
                    if normalize(posname) in skip_set:
                        continue

                bm = best_match_for_step(posname, df, step, is_spread)
                if bm is None:
                    continue

                tpl = bm["template"]
                phs = tpl["Placeholders"]
                choice = bm["choice"]
                pool_use = bm["pool_use"]

                selected=[]
                for slot,(idx,kind) in enumerate(choice):
                    if idx is None:
                        continue
                    prod = pool_use.loc[idx].to_dict()
                    selected.append((normalize(prod.get("Naam productie","")), phs[slot]))

                names_real=[n for n,_ph in selected]

                # Placeholder(s) die open blijven (idx=None)
                open_placeholders=[phs[slot] for slot,(idx,kind) in enumerate(choice) if idx is None]

                df_new = df[df["Naam productie"].astype(str).str.strip().isin(names_real)].copy()
                for col in ["Gekozen template","Gekozen placeholder","Plaatsing"]:
                    if col not in df_new.columns:
                        df_new[col]=None

                for i,row in df_new.iterrows():
                    nm = normalize(row["Naam productie"])
                    df_new.at[i,"Gekozen template"] = tpl["Template"]
                    df_new.at[i,"Gekozen placeholder"] = next((ph for n,ph in selected if n==nm), "")
                    df_new.at[i,"Plaatsing"] = posname

                # Voeg extra rij(en) toe als er placeholder(s) open blijven
                if open_placeholders:
                    # Safety: voorkom dubbele kolomnamen (pandas concat vereist unieke columns)
                    if df_new.columns.duplicated().any():
                        df_new = df_new.loc[:, ~df_new.columns.duplicated()].copy()
                    for oph in open_placeholders:
                        empty_row={c:None for c in df_new.columns}
                        empty_row["Naam productie"]="NOG ARTIKEL VOOR DEZE PLEK ZOEKEN"
                        empty_row["Gekozen template"]=tpl["Template"]
                        empty_row["Gekozen placeholder"]=oph
                        empty_row["Plaatsing"]=posname
                        df_new = pd.concat([df_new, pd.DataFrame([empty_row])], ignore_index=True)

                df_to_sheet(ws, df_new.reset_index(drop=True))


                # Copy / post-processing (standaard vs. UITW-runs)
                name_to_vals={}
                headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
                name_col = None
                for i,h in enumerate(headers, start=1):
                    if normalize(h).lower() == "naam productie":
                        name_col = i
                        break
                if name_col is None:
                    name_col = 1  # fallback: oude structuur
                for rr in range(2, ws.max_row+1):
                    nm = normalize(ws.cell(rr, name_col).value)
                    if nm:
                        name_to_vals[nm] = {headers[c-1]: ws.cell(rr,c).value for c in range(1, ws.max_column+1)}

                is_uitw_run = ("-U" in posname)

                if not is_uitw_run:
                    # Standaard: kopieer naar targets uit AA1 en overschrijf daar de rij
                    for tr in copy_targets:
                        if tr in plan_wb.sheetnames:
                            wst = plan_wb[tr]
                            for nm,_ph in selected:
                                overwrite_row_by_name(wst, nm, name_to_vals[nm])

                    # Verwijder gematchte producties van alle andere tabbladen behalve keep_targets (AB1)
                    exclude=set([posname]+keep_targets)
                    for sh in plan_wb.sheetnames:
                        if sh in exclude:
                            continue
                        wso = plan_wb[sh]
                        for nm in names_real:
                            remove_from_sheet(wso, nm)

                else:
                    # UITW-runs: append naar Totale verhalenlijst en strip alleen binnen de UITW-groep (NM-U* of ZU-U*)
                    if "Totale verhalenlijst" in plan_wb.sheetnames:
                        wtot = plan_wb["Totale verhalenlijst"]
                        tot_headers = [wtot.cell(1,c).value for c in range(1, wtot.max_column+1)]
                        tot_h2c = {str(h).strip(): i+1 for i,h in enumerate(tot_headers) if h is not None and str(h).strip() != ""}

                        for nm,_ph in selected:
                            vals = name_to_vals.get(nm, {})
                            new_row_idx = wtot.max_row + 1
                            for h, col in tot_h2c.items():
                                if h in vals:
                                    wtot.cell(new_row_idx, col).value = vals[h]

                    # Strip gematchte producties uit andere tabbladen
                    if posname in ("NM-U1","NM-U2","NM-U3","NM-U4","NM-U5","NM-U6","ZU-U1","ZU-U2","ZU-U3","ZU-U4","ZU-U5","ZU-U6"):
                        # volgens spec: strip uit alle andere tabbladen met dezelfde 2 beginletters ("NM")
                        prefix_main = posname[:2] + "-"
                        for sh in plan_wb.sheetnames:
                            if sh == posname or sh == "Totale verhalenlijst":
                                continue
                            if not str(sh).startswith(prefix_main):
                                continue
                            wso = plan_wb[sh]
                            for nm in names_real:
                                remove_from_sheet(wso, nm)
                    else:
                        # default (andere -U runs): strip alleen binnen de UITW-groep (NM-U* of ZU-U*)
                        prefix_u = posname[:2] + "-U"  # "NM-U" of "ZU-U"
                        for sh in plan_wb.sheetnames:
                            if sh == posname or sh == "Totale verhalenlijst":
                                continue
                            if not str(sh).startswith(prefix_u):
                                continue
                            wso = plan_wb[sh]
                            for nm in names_real:
                                remove_from_sheet(wso, nm)

                log(f"Run {posname}: stap {step['Stappen']} succesvolle match. Template={tpl['Template']}. {bm['conct']}")
                matched=True
                success += 1
                break

            if not matched:
                log(f"Run {posname}: {beslis.iloc[0]['Stappen']} tot en met {beslis.iloc[-1]['Stappen']} geen match.")



    # Volgens Opzet Krantenplanner DL (versimpeld):
    # Voer achtereenvolgens 7 runs uit: volgorde uit Planningsvolgorde A2:A8
    # (in het planningsbestand: tab 'Planningsvolgorde' -> cellen A2 t/m A8)
    runs = order[:7] if len(order) >= 7 else order

    run_runs(runs)



    print(f"Klaar. Succesvolle matches: {success}")

    # -----------------------------
    # Save output
    # -----------------------------
    # -----------------------------
    # EXTRA STAP (versie 5):
    # Maak kopieën van 'Totale verhalenlijst' -> 'ZU-UITW' en 'NM-UITW'
    # en zet Classificatie overal op 'A-keus; B-keus; C-keus;'
    # -----------------------------
    source_name = "Totale verhalenlijst"
    if source_name in plan_wb.sheetnames:
        # verwijder bestaande kopieën
        for nm in ["ZU-UITW", "NM-UITW"]:
            if nm in plan_wb.sheetnames:
                plan_wb.remove(plan_wb[nm])

        src_ws = plan_wb[source_name]

        def _copy_and_set_classificatie(new_title):
            ws_copy = plan_wb.copy_worksheet(src_ws)
            ws_copy.title = new_title

            headers = [ws_copy.cell(1,c).value for c in range(1, ws_copy.max_column+1)]
            class_col = None
            for i,h in enumerate(headers, start=1):
                if ("" if h is None else str(h).strip().lower()) == "classificatie":
                    class_col = i
                    break
            if class_col is None:
                class_col = ws_copy.max_column + 1
                ws_copy.cell(1, class_col).value = "Classificatie"

            # Nieuwe kolom Z: 'Telt mee' (overal 'ja')
            teltmee_col = 26  # kolom Z
            ws_copy.cell(1, teltmee_col).value = "Telt mee"

            for r in range(2, ws_copy.max_row+1):
                ws_copy.cell(r, class_col).value = "A-keus; B-keus; C-keus;"
                ws_copy.cell(r, teltmee_col).value = "ja"

        _copy_and_set_classificatie("ZU-UITW")
        _copy_and_set_classificatie("NM-UITW")


    # -----------------------------
    # EXTRA STAP (versie 6):
    # Volg [MAPPINGREGELS ZU-UITW]
    # -----------------------------
    def _find_df_col(df, name):
        tgt = normalize(name).lower()
        for c in df.columns:
            if normalize(c).lower() == tgt:
                return c
        return None

    def apply_mappingregels_zu_uitw():
        if "ZU-UITW" not in plan_wb.sheetnames:
            return
        ws = plan_wb["ZU-UITW"]
        df = sheet_to_df(ws)
        if df is None or df.empty:
            return

        # kolommen (robust)
        col_pl = _find_df_col(df, "Plaatsing")
        col_hl = _find_df_col(df, "Heel Limburg")
        col_fr = _find_df_col(df, "Focusregio")
        col_gp = _find_df_col(df, "Gewenste placeholder")
        col_ph_es = _find_df_col(df, "Placeholder bij enigszins geschikt")
        col_pr = _find_df_col(df, "Prioscore")
        col_gt = _find_df_col(df, "Gekozen template")
        col_gph= _find_df_col(df, "Gekozen placeholder")
        col_pd = _find_df_col(df, "Publicatiedwang")
        col_t8 = _find_df_col(df, "Top 8")
        col_tm = _find_df_col(df, "Telt mee")


        # zorg dat kolom "Telt mee" bestaat (default ja)
        if col_tm is None:
            df["Telt mee"] = "ja"
            col_tm = _find_df_col(df, "Telt mee")
        # zorg dat reset-kolommen bestaan
        for cname in ["Prioscore","Gekozen template","Gekozen placeholder","Plaatsing"]:
            if _find_df_col(df, cname) is None:
                df[cname] = None
        col_pr = _find_df_col(df, "Prioscore")
        col_gt = _find_df_col(df, "Gekozen template")
        col_gph= _find_df_col(df, "Gekozen placeholder")
        col_pl = _find_df_col(df, "Plaatsing")

        # 1) Schrap: Plaatsing begint met 'GO-' of 'ZU-'
        if col_pl is not None:
            pl = df[col_pl].fillna("").astype(str)
            df = df[~(pl.str.startswith("GO-") | pl.str.startswith("ZU-"))].copy()

        # 2) Schrap: Heel Limburg=ongeschikt AND Focusregio bevat géén Parkstad/Maastricht/Sittard
        targets = ["Parkstad","Maastricht","Sittard"]
        def _contains_any(txt):
            s = "" if txt is None else str(txt)
            return any(t in s for t in targets)

        if col_hl is not None:
            hl = df[col_hl].fillna("").astype(str).str.strip().str.lower()
            fr = df[col_fr].fillna("").astype(str) if col_fr is not None else pd.Series([""]*len(df), index=df.index)
            mask_bad = (hl == "ongeschikt") & (~fr.apply(_contains_any))
            df = df[~mask_bad].copy()

        # 3) Enigszins geschikt + Focusregio ≠ Sittard/Parkstad/Maastricht -> vervang Gewenste placeholder
        if col_hl is not None and col_fr is not None and col_gp is not None and col_ph_es is not None:
            hl = df[col_hl].fillna("").astype(str).str.strip().str.lower()
            fr = df[col_fr].fillna("").astype(str).str.strip()
            mask = (hl == "enigszins geschikt") & (~fr.isin(targets))
            df.loc[mask, col_gp] = df.loc[mask, col_ph_es]
            if col_tm is not None:
                df.loc[mask, col_tm] = "nee"

        # 4) Wis kolommen Prioscore, Gekozen template, Gekozen placeholder, Plaatsing
        df[col_pr] = None
        df[col_gt] = None
        df[col_gph] = None
        df[col_pl] = None

        # 5) Bereken nieuwe Prioscore
        score = pd.Series([0]*len(df), index=df.index, dtype="int64")

        if col_hl is not None:
            hl = df[col_hl].fillna("").astype(str).str.strip().str.lower()
            score += (hl == "moet mee").astype(int) * 5
            score += (hl == "geschikt").astype(int) * 3
            score += (hl == "enigszins geschikt").astype(int) * 1

        if col_pd is not None:
            pdw = df[col_pd].fillna("").astype(str).str.strip().str.lower()
            score += (pdw == "nee").astype(int) * (-1)

        if col_t8 is not None:
            t8 = df[col_t8].fillna("").astype(str).str.strip().str.lower()
            score += (t8 == "ja").astype(int) * 3

        if col_fr is not None:
            fr = df[col_fr].fillna("").astype(str)
            score += fr.apply(lambda x: 10 if _contains_any(x) else 0).astype(int)

        df[col_pr] = score

        # 6) Sorteer op Prioscore desc
        df = df.sort_values(by=[col_pr], ascending=False, kind="mergesort").reset_index(drop=True)

        df_to_sheet(ws, df)

    # run mappingregels
    apply_mappingregels_zu_uitw()


    # -----------------------------
    # EXTRA STAP (versie 7):
    # Volg [MAPPINGREGELS NM-UITW]
    # -----------------------------
    def apply_mappingregels_nm_uitw():
        if "NM-UITW" not in plan_wb.sheetnames:
            return
        ws = plan_wb["NM-UITW"]
        df = sheet_to_df(ws)
        if df is None or df.empty:
            return

        # kolommen (robust)
        col_pl = _find_df_col(df, "Plaatsing")
        col_hl = _find_df_col(df, "Heel Limburg")
        col_fr = _find_df_col(df, "Focusregio")
        col_gp = _find_df_col(df, "Gewenste placeholder")
        col_ph_es = _find_df_col(df, "Placeholder bij enigszins geschikt")
        col_pr = _find_df_col(df, "Prioscore")
        col_gt = _find_df_col(df, "Gekozen template")
        col_gph= _find_df_col(df, "Gekozen placeholder")
        col_pd = _find_df_col(df, "Publicatiedwang")
        col_t8 = _find_df_col(df, "Top 8")
        col_tm = _find_df_col(df, "Telt mee")


        # zorg dat kolom "Telt mee" bestaat (default ja)
        if col_tm is None:
            df["Telt mee"] = "ja"
            col_tm = _find_df_col(df, "Telt mee")
        # zorg dat reset-kolommen bestaan
        for cname in ["Prioscore","Gekozen template","Gekozen placeholder","Plaatsing"]:
            if _find_df_col(df, cname) is None:
                df[cname] = None
        col_pr = _find_df_col(df, "Prioscore")
        col_gt = _find_df_col(df, "Gekozen template")
        col_gph= _find_df_col(df, "Gekozen placeholder")
        col_pl = _find_df_col(df, "Plaatsing")

            # 1) Schrap: Plaatsing begint met 'GO-' of 'NM-'
        if col_pl is not None:
            pl = df[col_pl].fillna("").astype(str)
            df = df[~(pl.str.startswith("GO-") | pl.str.startswith("NM-"))].copy()

        # 2) Schrap: Heel Limburg=ongeschikt AND Focusregio bevat géén Noord/Midden
        targets = ["Noord","Midden"]
        def _contains_any_nm(txt):
            s = "" if txt is None else str(txt)
            return any(t in s for t in targets)

        if col_hl is not None:
            hl = df[col_hl].fillna("").astype(str).str.strip().str.lower()
            fr = df[col_fr].fillna("").astype(str) if col_fr is not None else pd.Series([""]*len(df), index=df.index)
            mask_bad = (hl == "ongeschikt") & (~fr.apply(_contains_any_nm))
            df = df[~mask_bad].copy()

        # 3) Enigszins geschikt + Focusregio ≠ Noord/Midden -> vervang Gewenste placeholder
        if col_hl is not None and col_fr is not None and col_gp is not None and col_ph_es is not None:
            hl = df[col_hl].fillna("").astype(str).str.strip().str.lower()
            fr = df[col_fr].fillna("").astype(str).str.strip()
            mask = (hl == "enigszins geschikt") & (~fr.isin(targets))
            df.loc[mask, col_gp] = df.loc[mask, col_ph_es]
            if col_tm is not None:
                df.loc[mask, col_tm] = "nee"

        # 4) Wis kolommen Prioscore, Gekozen template, Gekozen placeholder, Plaatsing
        df[col_pr] = None
        df[col_gt] = None
        df[col_gph] = None
        df[col_pl] = None

        # 5) Bereken nieuwe Prioscore
        score = pd.Series([0]*len(df), index=df.index, dtype="int64")

        if col_hl is not None:
            hl = df[col_hl].fillna("").astype(str).str.strip().str.lower()
            score += (hl == "moet mee").astype(int) * 5
            score += (hl == "geschikt").astype(int) * 3
            score += (hl == "enigszins geschikt").astype(int) * 1

        if col_pd is not None:
            pdw = df[col_pd].fillna("").astype(str).str.strip().str.lower()
            score += (pdw == "nee").astype(int) * (-1)

        if col_t8 is not None:
            t8 = df[col_t8].fillna("").astype(str).str.strip().str.lower()
            score += (t8 == "ja").astype(int) * 3

        if col_fr is not None:
            fr = df[col_fr].fillna("").astype(str)
            score += fr.apply(lambda x: 10 if _contains_any_nm(x) else 0).astype(int)

        df[col_pr] = score

        # 6) Sorteer op Prioscore desc
        df = df.sort_values(by=[col_pr], ascending=False, kind="mergesort").reset_index(drop=True)

        df_to_sheet(ws, df)

    # run mappingregels
    apply_mappingregels_nm_uitw()


    # -----------------------------
    # EXTRA STAP (versie 8):
    # Kopieer ZU-UITW -> ZU-U1..ZU-U6 en ZU-UNUSED, verwijder daarna ZU-UITW
    # Kopieer NM-UITW -> NM-U1..NM-U6 en NM-UNUSED, verwijder daarna NM-UITW
    # -----------------------------
    def _ensure_fresh_copy(src_name: str, dest_name: str):
        if dest_name in plan_wb.sheetnames:
            plan_wb.remove(plan_wb[dest_name])
        base_ws = plan_wb[src_name]
        new_ws = plan_wb.copy_worksheet(base_ws)
        new_ws.title = dest_name
        return new_ws

    def _copy_uitw_to_u_tabs(prefix: str):
        src = f"{prefix}-UITW"
        if src not in plan_wb.sheetnames:
            return
        # maak kopieën
        for i in range(1, 7):
            _ensure_fresh_copy(src, f"{prefix}-U{i}")
        _ensure_fresh_copy(src, f"{prefix}-UNUSED")
        # verwijder bron
        plan_wb.remove(plan_wb[src])

    _copy_uitw_to_u_tabs("ZU")
    _copy_uitw_to_u_tabs("NM")


    # -----------------------------
    # 7) Volg [MAPPINGREGELS EXTRA 1]
    # -----------------------------
    def _pos_header_map(ws):
        hdr = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
        return {("" if h is None else str(h).strip().lower()): i+1 for i,h in enumerate(hdr)}

    def _get_pos_col(ws, header_name: str):
        m = _pos_header_map(ws)
        return m.get(header_name.strip().lower())

    def _apply_extra1_for_tab(tabname: str, bonus: int, append_twede: bool):
        if tabname not in plan_wb.sheetnames:
            return
        ws = plan_wb[tabname]
        # header -> col index
        hmap = _pos_header_map(ws)
        def col(*names):
            for n in names:
                c = hmap.get(n.strip().lower())
                if c:
                    return c
            return None

        col_hl   = col("Heel Limburg")
        col_prio = col("Prioscore")
        col_gw   = col("Gewenste placeholder")
        col_tw   = col("Tweede keus placeholder")

        if col_hl is None or col_prio is None:
            return  # kan niets doen zonder deze kolommen

        for r in range(2, ws.max_row+1):
            hl = normalize(ws.cell(r, col_hl).value).strip().lower()
            if hl != "moet mee":
                continue

            # Prioscore +bonus
            cur = ws.cell(r, col_prio).value
            try:
                cur_f = float(cur) if cur not in (None, "") else 0.0
            except Exception:
                cur_f = 0.0
            ws.cell(r, col_prio).value = cur_f + bonus

            # optioneel: Tweede keus placeholder toevoegen aan Gewenste placeholder
            if append_twede and col_gw is not None and col_tw is not None:
                gw = "" if ws.cell(r, col_gw).value is None else str(ws.cell(r, col_gw).value).strip()
                tw = "" if ws.cell(r, col_tw).value is None else str(ws.cell(r, col_tw).value).strip()
                if tw:
                    # voeg toe als het nog niet in de string zit (case-sensitive behouden)
                    if not gw:
                        ws.cell(r, col_gw).value = tw
                    elif tw not in gw:
                        sep = "; " if not gw.rstrip().endswith(";") else " "
                        ws.cell(r, col_gw).value = gw + sep + tw

    # bepaal tabbladen op basis van [POSITIELIJST] kolom 'Restant'
    rest_col = None
    # pos_header en pos_ws bestaan al eerder in het notebook
    try:
        rest_col = None
        # gebruik bestaande pos_col helper als die bestaat
        if "pos_col" in globals():
            try:
                rest_col = pos_col("Restant")
            except Exception:
                rest_col = None
        if rest_col is None:
            hdr = [pos_ws.cell(1,c).value for c in range(1, pos_ws.max_column+1)]
            for i,h in enumerate(hdr, start=1):
                if normalize(h).strip().lower() == "restant":
                    rest_col = i
                    break
    except Exception:
        rest_col = None

    if rest_col is not None:
        for rr in range(2, pos_ws.max_row+1):
            tab = normalize(pos_ws.cell(rr, POS_POS_COL).value)
            if not tab:
                continue
            val = pos_ws.cell(rr, rest_col).value
            try:
                v = int(float(val)) if val not in (None, "") else None
            except Exception:
                v = None
            if v in (3,4):
                _apply_extra1_for_tab(tab, bonus=20, append_twede=False)
            elif v in (1,2):
                _apply_extra1_for_tab(tab, bonus=30, append_twede=True)



    # EXTRA 1 - stap 3 t/m 7: specifieke bewerkingen op ZU-U1, NM-U1, ZU-U2, NM-U2 + opnieuw sorteren
    def _remove_tokens(s, tokens):
        if s is None:
            return s
        txt = str(s)
        for tok in tokens:
            # verwijder token met evt. whitespace eromheen, behoud reststring
            txt = re.sub(rf"\b{re.escape(tok)}\s*;\s*", "", txt)
        txt = re.sub(r"\s{2,}", " ", txt).strip()
        return txt

    def _apply_extra1_tab(tabname: str, focus_remove_list, focus_bonus_list, rules):
        """
        rules: list van dicts met keys:
          - hl_values: tuple van hl-waarden (lowercase) waarop regel geldt
          - remove_tokens: lijst tokens om te verwijderen uit Classificatie (bij focus NOT contains focus_remove_list)
        """
        if tabname not in plan_wb.sheetnames:
            return
        ws = plan_wb[tabname]
        df = sheet_to_df(ws)
        if df is None or df.empty:
            return

        c_hl  = _find_df_col(df, "Heel Limburg")
        c_fr  = _find_df_col(df, "Focusregio")
        c_cl  = _find_df_col(df, "Classificatie")
        c_pr  = _find_df_col(df, "Prioscore")

        if c_hl is None or c_fr is None or c_pr is None or c_cl is None:
            return

        def contains_any(text, needles):
            if text is None:
                return False
            t = str(text)
            return any(n in t for n in needles)

        for i in range(len(df)):
            hl = "" if df.at[i, c_hl] is None else str(df.at[i, c_hl]).strip().lower()
            fr = "" if df.at[i, c_fr] is None else str(df.at[i, c_fr])

            # verwijder tokens afhankelijk van HL, maar alleen als focusregio géén van de targets bevat
            if not contains_any(fr, focus_remove_list):
                for rule in rules:
                    if hl in rule["hl_values"]:
                        df.at[i, c_cl] = _remove_tokens(df.at[i, c_cl], rule["remove_tokens"])

            # +5 bij focusregio match (bonuslijst)
            if contains_any(fr, focus_bonus_list):
                try:
                    cur = float(df.at[i, c_pr]) if df.at[i, c_pr] not in (None, "") else 0.0
                except Exception:
                    cur = 0.0
                df.at[i, c_pr] = cur + 5

        # sorteer op Prioscore (hoogste bovenaan)
        def prio_key(x):
            try:
                return float(x)
            except Exception:
                return -1e18
        df["_prio_sort"] = df[c_pr].apply(prio_key)
        df = df.sort_values("_prio_sort", ascending=False).drop(columns=["_prio_sort"]).reset_index(drop=True)
        df_to_sheet(ws, df)

    # 3) ZU-U1
    _apply_extra1_tab(
        "ZU-U1",
        focus_remove_list=["Parkstad","Maastricht","Sittard","Limburg-breed"],
        focus_bonus_list=["Parkstad","Maastricht","Sittard"],
        rules=[
            {"hl_values": ("geschikt",), "remove_tokens": ["A-keus"]},
            {"hl_values": ("enigszins geschikt",), "remove_tokens": ["A-keus","B-keus"]},
        ],
    )

    # 4) NM-U1
    _apply_extra1_tab(
        "NM-U1",
        focus_remove_list=["Noord","Midden","Limburg-breed"],
        focus_bonus_list=["Noord","Midden"],
        rules=[
            {"hl_values": ("geschikt",), "remove_tokens": ["A-keus"]},
            {"hl_values": ("enigszins geschikt",), "remove_tokens": ["A-keus","B-keus"]},
        ],
    )

    # 5) ZU-U2
    _apply_extra1_tab(
        "ZU-U2",
        focus_remove_list=["Parkstad","Maastricht","Sittard","Limburg-breed"],
        focus_bonus_list=["Parkstad","Maastricht","Sittard","Limburg-breed"],
        rules=[
            {"hl_values": ("enigszins geschikt",), "remove_tokens": ["A-keus"]},
        ],
    )

    # 6) NM-U2
    _apply_extra1_tab(
        "NM-U2",
        focus_remove_list=["Noord","Midden","Limburg-breed"],
        focus_bonus_list=["Noord","Midden","Limburg-breed"],
        rules=[
            {"hl_values": ("enigszins geschikt",), "remove_tokens": ["A-keus"]},
        ],
    )



    # 7) Rangschik alle U-tabbladen op Prioscore (hoogste bovenaan) — geldt ook voor NM-U3 e.v.
    def _sort_sheet_by_prioscore(tabname: str):
        if tabname not in plan_wb.sheetnames:
            return
        ws = plan_wb[tabname]
        df = sheet_to_df(ws)
        if df is None or df.empty:
            return
        c_pr = _find_df_col(df, "Prioscore")
        if c_pr is None:
            return

        def prio_key(x):
            try:
                return float(x)
            except Exception:
                return -1e18

        df["_prio_sort"] = df[c_pr].apply(prio_key)
        df = df.sort_values("_prio_sort", ascending=False, kind="mergesort").drop(columns=["_prio_sort"]).reset_index(drop=True)
        df_to_sheet(ws, df)

    for sh in list(plan_wb.sheetnames):
        if re.match(r"^(NM|ZU)-U(\d+|UNUSED)$", sh):
            _sort_sheet_by_prioscore(sh)

    # 8) Voer de volgende run uit: NM-U1
    run_runs(["NM-U1","NM-U2","ZU-U1","ZU-U2"]) 


    # 9) Volg [OPSPAARREGELS]
    def _ws_find_col(ws, header_name: str):
        tgt = ("" if header_name is None else str(header_name)).strip().lower()
        for c in range(1, ws.max_column+1):
            h = ws.cell(1,c).value
            if ("" if h is None else str(h)).strip().lower() == tgt:
                return c
        return None

    def _delete_story_id_rows(ws, story_col, target_story_id):
        # delete bottom-up to avoid skipping
        for r in range(ws.max_row, 1, -1):
            v = ws.cell(r, story_col).value
            if v is None:
                continue
            if str(v).strip() == target_story_id:
                ws.delete_rows(r, 1)

    if "Totale verhalenlijst" in plan_wb.sheetnames:
        ws_tot = plan_wb["Totale verhalenlijst"]
        col_pd = _ws_find_col(ws_tot, "Publicatiedwang")
        col_pl = _ws_find_col(ws_tot, "Plaatsing")
        col_sid = _ws_find_col(ws_tot, "Story ID")

        if col_pd is not None and col_sid is not None:
            # tel Story ID voorkomen in Totale verhalenlijst
            sid_counts = {}
            for r in range(2, ws_tot.max_row+1):
                sid = ws_tot.cell(r, col_sid).value
                if sid is None or str(sid).strip() == "":
                    continue
                key = str(sid).strip()
                sid_counts[key] = sid_counts.get(key, 0) + 1

            # bepaal welke story ids 'op te sparen' zijn en (nog) geen plaatsing hebben
            target_sids = set()
            for r in range(2, ws_tot.max_row+1):
                pdw = ws_tot.cell(r, col_pd).value
                if normalize(pdw).strip().lower() != "op te sparen":
                    continue
                plv = ws_tot.cell(r, col_pl).value if col_pl is not None else None
                if plv is not None and str(plv).strip() != "":
                    continue  # Plaatsing gevuld -> niet meenemen
                sid = ws_tot.cell(r, col_sid).value
                if sid is None or str(sid).strip() == "":
                    continue
                key = str(sid).strip()
                # alleen actie als story id uniek is op Totale verhalenlijst
                if sid_counts.get(key, 0) == 1:
                    target_sids.add(key)

            # schrap op U3..U6 tabbladen de rijen met dezelfde Story ID (alleen als uniek op Totale verhalenlijst)
            target_tabs = ["ZU-U3","ZU-U4","ZU-U5","ZU-U6","NM-U3","NM-U4","NM-U5","NM-U6"]
            for tab in target_tabs:
                if tab not in plan_wb.sheetnames:
                    continue
                ws = plan_wb[tab]
                c_sid = _ws_find_col(ws, "Story ID")
                if c_sid is None:
                    continue
                for sid in target_sids:
                    _delete_story_id_rows(ws, c_sid, sid)



    # 10) Voer de volgende runs uit: NM-U3, NM-U4, NM-U5, NM-U6, ZU-U3, ZU-U4, ZU-U5, ZU-U6.
    run_runs(["NM-U3","NM-U4","NM-U5","NM-U6","ZU-U3","ZU-U4","ZU-U5","ZU-U6"])



    # 11) Verwijder op 'Totale verhalenlijst' alle rijen waar kolom 'Plaatsing' leeg is.
    #     Sorteer daarna 'Totale verhalenlijst' op 'Naam productie'.
    def _cleanup_totale_verhalenlijst():
        sh = "Totale verhalenlijst"
        if sh not in plan_wb.sheetnames:
            return
        ws = plan_wb[sh]
        df = sheet_to_df(ws)
        if df is None or df.empty:
            return

        c_pl = _find_df_col(df, "Plaatsing")
        c_nm = _find_df_col(df, "Naam productie")
        if c_pl is None:
            return

        pl = df[c_pl].fillna("").astype(str).str.strip()
        df = df[pl != ""].copy()

        if c_nm is not None:
            df = df.sort_values(by=[c_nm], ascending=True, kind="mergesort").reset_index(drop=True)

        df_to_sheet(ws, df)

    _cleanup_totale_verhalenlijst()

    # 12) Kopieer in alle tabbladen waarvoor een run is uitgevoerd de bijbehorende 'Beschrijving' uit Logfile
    #     naar cel AD1 van dat tabblad (dus per tabblad de juiste regel, niet de laatste algemene regel).
    def _copy_log_beschrijving_to_tabs():
        if "Logfile" not in plan_wb.sheetnames:
            return
        ws_log = plan_wb["Logfile"]

        # zoek kolom 'Beschrijving' op rij 1
        desc_col = None
        for c in range(1, ws_log.max_column + 1):
            if normalize(ws_log.cell(1, c).value).strip().lower() == "beschrijving":
                desc_col = c
                break
        if desc_col is None:
            return

        # tabs waarvoor runs zijn uitgevoerd: Planningsvolgorde + alle U-tabs + GO-01/GO-02 (als aanwezig)
        tabs = []
        seen = set()
        try:
            for t in order:
                if t and str(t) not in seen:
                    tabs.append(str(t)); seen.add(str(t))
        except Exception:
            pass
        for sh in plan_wb.sheetnames:
            shs = str(sh)
            if re.match(r"^(NM|ZU)-U(\d+|UNUSED)$", shs) and shs not in seen:
                tabs.append(shs); seen.add(shs)
        for t in ["GO-01", "GO-02"]:
            if t in plan_wb.sheetnames and t not in seen:
                tabs.append(t); seen.add(t)

        # bouw per tab de laatst voorkomende beschrijving: scan Logfile van onder naar boven
        last_for_tab = {t: None for t in tabs}
        # descriptions beginnen in Logfile met 'Run <TAB>:'
        targets = {t: f"Run {t}:" for t in tabs}

        for r in range(ws_log.max_row, 1, -1):
            v = ws_log.cell(r, desc_col).value
            if v in (None, ""):
                continue
            s = str(v)
            for t, prefix in targets.items():
                if last_for_tab[t] is None and s.startswith(prefix):
                    last_for_tab[t] = v
            # early exit als alles gevonden
            if all(last_for_tab[t] is not None for t in tabs):
                break

        # schrijf naar AD1
        for t in tabs:
            if t in plan_wb.sheetnames and last_for_tab.get(t) is not None:
                plan_wb[t]["AD1"].value = last_for_tab[t]

    _copy_log_beschrijving_to_tabs()



    # -----------------------------
    # 13) Voer [EINDBEWERKINGEN] uit
    # -----------------------------
    def _find_ws_col(ws, header_name: str):
        target = str(header_name).strip().lower()
        for c in range(1, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h is None:
                continue
            if str(h).strip().lower() == target:
                return c
        return None

    def _ensure_ws_col(ws, header_name: str):
        col = _find_ws_col(ws, header_name)
        if col is not None:
            return col
        col = ws.max_column + 1
        ws.cell(1, col).value = header_name
        return col

    def _get_run_tabs_from_logfile():
        if "Logfile" not in plan_wb.sheetnames:
            return []
        ws = plan_wb["Logfile"]
        # zoek kolom 'Beschrijving'
        desc_col = _find_ws_col(ws, "Beschrijving")
        if desc_col is None:
            return []
        tabs = []
        for r in range(2, ws.max_row + 1):
            txt = ws.cell(r, desc_col).value
            if not txt:
                continue
            s = str(txt)
            # verwacht: "Run <TAB> uitgevoerd, ..."
            m = re.match(r"\s*Run\s+([A-Za-z]{2}-[A-Za-z0-9]+)\s+uitgevoerd", s)
            if not m:
                # fallback: "Run <TAB>: ..."
                m = re.match(r"\s*Run\s+([A-Za-z]{2}-[A-Za-z0-9]+)\s*[:]", s)
            if m:
                tabs.append(m.group(1))
        # uniek, behoud volgorde
        seen=set()
        out=[]
        for t in tabs:
            if t not in seen:
                seen.add(t); out.append(t)
        return out

    def _apply_placeholder_concessie(ws):
        col_gp = _find_ws_col(ws, "Gekozen placeholder")
        col_wp = _find_ws_col(ws, "Gewenste placeholder")
        if col_gp is None or col_wp is None:
            return
        col_pc = _ensure_ws_col(ws, "Placeholder-concessie")
        for r in range(2, ws.max_row + 1):
            gp = ws.cell(r, col_gp).value
            wp = ws.cell(r, col_wp).value
            gp_s = "" if gp is None else str(gp).strip()
            wp_s = "" if wp is None else str(wp)
            if gp_s and (gp_s not in wp_s):
                ws.cell(r, col_pc).value = "Ja"
            else:
                # laat leeg (geen waarde gespecificeerd in txt)
                ws.cell(r, col_pc).value = None

    def _write_naam_van_positie_to_AE1(tabname: str):
        if tabname not in plan_wb.sheetnames:
            return
        # zoek kolom 'Naam van positie' in POSITIELIJST
        nv_col = None
        for i,h in enumerate(pos_header, start=1):
            if normalize(h).strip().lower() == "naam van positie":
                nv_col = i
                break
        if nv_col is None:
            return
        pr = get_pos_row(tabname)
        if pr is None:
            return
        val = pos_ws.cell(pr, nv_col).value
        if val is None:
            return
        plan_wb[tabname]["AE1"].value = val

    # 13.1 Placeholder-concessie op alle run-tabbladen + Totale verhalenlijst
    run_tabs = _get_run_tabs_from_logfile()
    for t in run_tabs:
        if t in plan_wb.sheetnames:
            _apply_placeholder_concessie(plan_wb[t])
    # ook op Totale verhalenlijst
    if "Totale verhalenlijst" in plan_wb.sheetnames:
        _apply_placeholder_concessie(plan_wb["Totale verhalenlijst"])

    # 13.2 Vul op run-tabbladen cel AE1 met 'Naam van positie' uit POSITIELIJST
    for t in run_tabs:
        _write_naam_van_positie_to_AE1(t)




    # 13.3 Maak tabblad 'Planning print' als kopie van 'Totale verhalenlijst' en zet vooraan
    if "Totale verhalenlijst" in plan_wb.sheetnames:
        # verwijder bestaande Planning print indien aanwezig
        if "Planning print" in plan_wb.sheetnames:
            plan_wb.remove(plan_wb["Planning print"])
        ws_src = plan_wb["Totale verhalenlijst"]
        ws_pp = plan_wb.copy_worksheet(ws_src)
        ws_pp.title = "Planning print"
        # positioneer als eerste (meest links)
        try:
            plan_wb._sheets.remove(ws_pp)
            plan_wb._sheets.insert(0, ws_pp)
        except Exception:
            pass

        # 13.4 Verwijder kolommen op Planning print (op basis van headernaam)
        cols_to_remove = [
            "Note","Voorkeurspositie","Printbeeld","Graphic","Karakters","Artikelsoort",
            "Gewenste placeholder","Tweede keus placeholder","Derde keus placeholder","Vierde keus placeholder",
            "Placeholder bij enigszins geschikt","Prioscore"
        ]
        # bouw header->kolomindex map
        headers = [ws_pp.cell(1,c).value for c in range(1, ws_pp.max_column+1)]
        name_to_col = {}
        for i,h in enumerate(headers, start=1):
            if h is None:
                continue
            name_to_col[str(h).strip().lower()] = i
        # verzamel te verwijderen kolommen die bestaan
        cols_idx = []
        for nm in cols_to_remove:
            ci = name_to_col.get(nm.strip().lower())
            if ci is not None:
                cols_idx.append(ci)
        # verwijder van rechts naar links
        for ci in sorted(set(cols_idx), reverse=True):
            ws_pp.delete_cols(ci, 1)

        # 13.5 Sorteer rijen op basis van kolom 'Plaatsing' volgens vaste volgorde
        df_pp = sheet_to_df(ws_pp)
        if df_pp is not None and not df_pp.empty:
            c_pl = _find_df_col(df_pp, "Plaatsing")
            if c_pl is not None:
                order = [
                    "NM-NO","NM-MI","ZU-MH","ZU-SG","ZU-PS",
                    # legacy/alternatief: ND vs GO
                    "GO-01","GO-02","ND-01","ND-02",
                    "NM-U1","NM-U2","NM-U3","NM-U4","NM-U5","NM-U6",
                    "ZU-U1","ZU-U2","ZU-U3","ZU-U4","ZU-U5","ZU-U6"
                ]
                rank = {p:i for i,p in enumerate(order)}
                def _rk(x):
                    if x is None:
                        return 10**9
                    s = str(x).strip()
                    return rank.get(s, 10**9)
                df_pp["_pl_rank"] = df_pp[c_pl].apply(_rk)
                df_pp = df_pp.sort_values(by=["_pl_rank"], ascending=True, kind="mergesort").drop(columns=["_pl_rank"]).reset_index(drop=True)
                df_to_sheet(ws_pp, df_pp)



    # 13.6 Kopieer 'Totale verhalenlijst' naar 'ZU-VERV' en 'NM-VERV' (en 'GO-VERV' alleen als er een GO-tabblad actief is)
    def _copy_sheet(src_name: str, dest_name: str, make_first: bool=False):
        if src_name not in plan_wb.sheetnames:
            return None
        if dest_name in plan_wb.sheetnames:
            plan_wb.remove(plan_wb[dest_name])
        ws_new = plan_wb.copy_worksheet(plan_wb[src_name])
        ws_new.title = dest_name
        if make_first:
            # zet meest links
            plan_wb._sheets.remove(ws_new)
            plan_wb._sheets.insert(0, ws_new)
        return ws_new

    _copy_sheet("Totale verhalenlijst", "ZU-VERV")
    _copy_sheet("Totale verhalenlijst", "NM-VERV")

    has_go_active = any(str(s).startswith("GO") for s in plan_wb.sheetnames)
    if has_go_active:
        _copy_sheet("Totale verhalenlijst", "GO-VERV")

    def _filter_verv_by_rules(tabname: str, prefix: str):
        if tabname not in plan_wb.sheetnames:
            return
        ws = plan_wb[tabname]
        df = sheet_to_df(ws)
        if df is None or df.empty:
            return

        c_pl = _find_df_col(df, "Plaatsing")
        c_fr = _find_df_col(df, "Focusregio")
        c_pd = _find_df_col(df, "Publicatiedwang")
        c_hl = _find_df_col(df, "Heel Limburg")

        # 13.7: strip rijen waarvan Plaatsing niet begint met prefix
        if c_pl is not None:
            pl = df[c_pl].fillna("").astype(str).str.strip()
            df = df[pl.str.startswith(prefix)].copy()

        # 13.8: extra stripregels
        def _norm(s):
            return normalize(s).strip().lower()

        if tabname == "NM-VERV":
            # Focusregio = Noord/Midden/Limburg-breed AND Publicatiedwang=Ja
            if c_fr is not None and c_pd is not None:
                fr = df[c_fr].fillna("").astype(str).str.strip()
                pdw = df[c_pd].fillna("").astype(str).apply(_norm)
                mask = fr.isin(["Noord","Midden","Limburg-breed"]) & (pdw == "ja")
                df = df[~mask].copy()
            # Heel Limburg = moet mee
            if c_hl is not None:
                hl = df[c_hl].fillna("").astype(str).apply(_norm)
                df = df[hl != "moet mee"].copy()

        elif tabname == "ZU-VERV":
            if c_fr is not None and c_pd is not None:
                fr = df[c_fr].fillna("").astype(str).str.strip()
                pdw = df[c_pd].fillna("").astype(str).apply(_norm)
                mask = fr.isin(["Parkstad","Sittard","Maastricht","Limburg-breed"]) & (pdw == "ja")
                df = df[~mask].copy()
            if c_hl is not None:
                hl = df[c_hl].fillna("").astype(str).apply(_norm)
                df = df[hl != "moet mee"].copy()

        elif tabname == "GO-VERV":
            if c_pd is not None:
                pdw = df[c_pd].fillna("").astype(str).apply(_norm)
                df = df[pdw != "ja"].copy()


        # 13.9: strip rijen waarvoor Top 8 = Ja (op alle *-VERV tabbladen)
        c_t8 = _find_df_col(df, "Top 8")
        if c_t8 is not None:
            t8 = df[c_t8].fillna("").astype(str).apply(_norm)
            df = df[t8 != "ja"].copy()

        df = df.reset_index(drop=True)
        df_to_sheet(ws, df)

    _filter_verv_by_rules("NM-VERV", "NM")
    _filter_verv_by_rules("ZU-VERV", "ZU")
    if has_go_active:
        _filter_verv_by_rules("GO-VERV", "GO")

    OUT_PATH = out_path
    plan_wb.save(OUT_PATH)
    return out_path
